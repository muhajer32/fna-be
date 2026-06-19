"""
main.py - Belgium FNA-ED/UC v2 unified workflow.

Runs either a deterministic single case or a Monte Carlo uncertainty run from
the same entry point. Excel/config controls the default mode; CLI flags can
override it for scripts and macros.
"""
from __future__ import annotations

import logging
import shutil
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook


from fna.config import (
    EXCEL_FILENAME,
    PATHS,
    PECD_COUNTRY_CODE,
    PECD_DATA_DIR,
    PECD_YEARS,
    PROJECT_ROOT,
    WIND_RESOURCE_IDS,
    make_run_paths,
    resolve_gms_path,
    output_excel_path,
    resolve_input_workbook,
    runs_root,
    csv_output_dir,
)
from fna.io.excel import (
    apply_uncertainty_scenario,
    parse_csv_results,
    read_inputs,
    read_inputs_with_mc,
    read_uncertainty_params,
    write_mc_charts_to_excel,
    write_mc_results_to_excel,
    write_results,
)
from fna.model.monte_carlo import MonteCarloScenarios, PECDReader, RepresentativeHourMap
from fna.plots.base import generate_all, generate_mc_summary_charts
from fna.model.gams import run_model

log = logging.getLogger("main")


# ---------------------------------------------------------------------------
# Run provenance / output isolation
# ---------------------------------------------------------------------------

_SETTINGS_KEYS = (
    "target_year", "future_year", "use_uc", "use_network", "soc_cyclic",
    "run_monte_carlo", "n_mc_scenarios", "max_parallel_workers", "seed_random",
    "use_pecd_data", "pecd_target_year", "market_time_unit_minutes",
    "curt_penalty", "reserve_slack_penalty", "network_slack_penalty", "voll",
    "entsoe_country_code", "entsoe_data_year",
)


def _settings_snapshot(inputs: dict[str, Any] | None, mc_params: dict[str, Any] | None = None) -> dict[str, Any]:
    """Capture the run-defining 01_Control values for the manifest, so a run's
    folder records the settings it was produced with (deterministic vs MC, year,
    switches, penalties, scenario count)."""

    control = dict((inputs or {}).get("control") or {})
    if mc_params:
        control.update({k: mc_params.get(k) for k in ("n_mc_scenarios", "max_parallel_workers", "seed_random", "use_pecd_data", "pecd_target_year") if k in mc_params})
    snapshot = {k: control.get(k) for k in _SETTINGS_KEYS if k in control}
    if inputs and inputs.get("target_year") is not None:
        snapshot.setdefault("target_year", inputs.get("target_year"))
    return snapshot


def begin_run(
    mode: str,
    inputs: dict[str, Any] | None,
    mc_params: dict[str, Any] | None = None,
    input_filename: str | None = None,
    n_scenarios: int | None = None,
) -> tuple[dict[str, Any], Any]:
    """Mint an isolated run directory + provenance manifest.

    Returns ``(paths, manifest)`` where ``paths`` is the run-rooted path dict
    (see ``config.make_run_paths``) and ``manifest`` is a started
    ``run_metadata.RunManifest``. Used by both ``main()`` and the CLI commands so
    every run lands in its own ``data/outputs/runs/<run_id>/`` folder."""

    from fna.run_metadata import RunManifest, capture_environment, new_run_id

    target_year = (inputs or {}).get("target_year")
    stem = Path(input_filename or EXCEL_FILENAME).stem
    run_id = new_run_id(mode, target_year, stem)
    paths = make_run_paths(run_id, input_filename)
    manifest = RunManifest(
        run_id=run_id,
        mode=mode,
        target_year=target_year,
        input_workbook=str(input_filename or EXCEL_FILENAME),
        run_dir=str(paths["run_dir"]),
        settings=_settings_snapshot(inputs, mc_params),
        environment=capture_environment(PROJECT_ROOT),
        n_scenarios_requested=n_scenarios,
    ).mark_started()
    log.info("Run id: %s", run_id)
    log.info("Run directory: %s", paths["run_dir"])
    return paths, manifest


def finish_run(manifest: Any, paths: dict[str, Any], out_wb: Any, status: str = "completed", error: str | None = None) -> None:
    """Finalise a run: stamp end time, write the metadata sheet(s) into the
    output workbook, and persist the JSON manifest + registry row. Best-effort
    - never raises into the run."""

    if manifest is None:
        return
    try:
        from fna.run_metadata import persist, write_metadata_sheets

        manifest.mark_finished(status=status, error=error)
        if out_wb is not None:
            write_metadata_sheets(out_wb, manifest)
        persist(manifest, runs_root())
    except Exception as exc:
        log.warning("Could not finalise run manifest: %s", exc)


def _output_path(paths: dict[str, Any] | None) -> Path:
    """Output workbook path for a run (per-run when isolated, else legacy)."""
    if paths and paths.get("output_xlsx"):
        return Path(paths["output_xlsx"])
    return output_excel_path()


def main() -> None:
    started_at = time.perf_counter()
    wb = _open_workbook()
    params = read_uncertainty_params(wb)
    run_mc = _use_monte_carlo(params)
    log_file = PROJECT_ROOT / "logs" / ("run_mc.log" if run_mc else "run.log")
    _configure_logging(log_file)

    try:
        if run_mc:
            run_monte_carlo(wb, params, started_at=started_at)
        else:
            run_optimisation(wb, started_at=started_at)
    finally:
        log.info("Total main.py runtime: %s", _format_elapsed(time.perf_counter() - started_at))


def run_optimisation(
    wb: Any | None = None,
    started_at: float | None = None,
    inputs: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
    manifest: Any | None = None,
) -> dict[str, Any]:
    """Run the deterministic mean/base case.

    ``inputs``, ``paths`` and ``sheet_suffix`` let multi_year.py reuse this
    function for each target year: ``inputs`` overrides the workbook-read
    inputs (e.g. with a year-specific target_year and demand column), ``paths``
    overrides the global PATHS with a per-year folder, and ``sheet_suffix`` is
    appended to every output sheet name. Returns the GAMS results dict, with
    the extra ACER tables (from write_results) under "fna_tables".

    When ``paths`` is None an isolated run folder is minted
    (``data/outputs/runs/<run_id>/``) and a provenance/compute manifest is
    recorded; pass ``paths`` (and optionally ``manifest``) to run inside an
    externally-managed folder, e.g. multi_year's per-year layout.
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 deterministic run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    if inputs is None:
        log.info("Reading v2 Excel inputs...")
        inputs = read_inputs(wb)

    owns_run = paths is None
    if owns_run:
        paths, manifest = begin_run("deterministic", inputs)
    paths["gms_file"] = resolve_gms_path(inputs.get("control"))
    out_wb = _open_output_workbook(out_path=_output_path(paths))
    status_cell = _safe_status_cell(wb)

    def status(msg: str) -> None:
        _set_status(status_cell, msg)
        log.info(msg)

    try:
        _ensure_directories(paths)

        status("Running GAMS 25.1 deterministic model...")
        results = run_model(inputs, {k: str(v) for k, v in paths.items()})
        _record_solo_timing(manifest, results)

        status("Writing output tables...")
        extra_tables = write_results(
            out_wb,
            results,
            rep_hours=inputs["frames"].get("hours"),
            rep_days=inputs["frames"].get("days"),
            network_needs=inputs["frames"].get("network"),
            dso_zones=inputs["frames"].get("dso_zones"),
            flex_availability=inputs["frames"].get("flex_profiles"),
            flex=inputs["frames"].get("flex"),
            prequalification=inputs["frames"].get("prequalification"),
            control=inputs.get("control"),
            target_year=inputs.get("target_year"),
            input_frames=inputs["frames"],
            sheet_suffix=sheet_suffix,
            img_dir=Path(paths["img_dir"]),
        )
        results["fna_tables"] = extra_tables

        status("Generating FNA output charts...")
        generate_all(results=results, img_dir=Path(paths["img_dir"]), wb=out_wb)

        if owns_run:
            finish_run(manifest, paths, out_wb, status="completed")
        if save:
            out_wb.save(_output_path(paths))
        status(f"Done: deterministic optimisation completed in {_elapsed_since(started_at)}.")
        if owns_run:
            status(f"Run outputs in {paths['run_dir']}")
        return results
    except Exception as exc:
        status("Error: see the run's gams_run.log / gams_run.lst")
        log.exception("Run failed: %s", exc)
        if owns_run:
            finish_run(manifest, paths, None, status="failed", error=str(exc))
        raise


def _record_solo_timing(manifest: Any, results: dict[str, Any]) -> None:
    """Record GAMS timing from a single (deterministic) solve onto the manifest."""
    if manifest is None:
        return
    timing = results.get("_gams_timing") or {}
    manifest.gams_wall_seconds_total = timing.get("gams_wall_seconds")


def run_postprocess(
    wb: Any | None = None,
    started_at: float | None = None,
    inputs: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
) -> dict[str, Any]:
    """Recompute the ACER FNA indicator / network-need sheets and charts from
    the CSV outputs of a *previous* GAMS run, without re-running GAMS.

    Used by the ``compute-fna-indicators`` and ``fine-tune`` CLI commands, so
    the (slow) optimisation does not have to be repeated just to re-derive the
    post-processed indicator tables.
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 post-processing run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    paths = paths or PATHS
    out_wb = _open_output_workbook(out_path=_output_path(paths))

    out_dir = Path(paths["out_dir"])
    log.info("Reading existing GAMS CSV outputs from %s", out_dir)
    results = parse_csv_results(out_dir)

    if inputs is None:
        log.info("Reading v2 Excel inputs...")
        inputs = read_inputs(wb)

    log.info("Writing FNA indicator / network-need tables...")
    extra_tables = write_results(
        out_wb,
        results,
        rep_hours=inputs["frames"].get("hours"),
        rep_days=inputs["frames"].get("days"),
        network_needs=inputs["frames"].get("network"),
        dso_zones=inputs["frames"].get("dso_zones"),
        flex_availability=inputs["frames"].get("flex_profiles"),
        flex=inputs["frames"].get("flex"),
        prequalification=inputs["frames"].get("prequalification"),
        control=inputs.get("control"),
        target_year=inputs.get("target_year"),
        input_frames=inputs["frames"],
        sheet_suffix=sheet_suffix,
        img_dir=Path(paths["img_dir"]),
    )
    results["fna_tables"] = _collect_indicator_tables(results, inputs, extra_tables)

    log.info("Generating FNA output charts...")
    generate_all(results=results, img_dir=Path(paths["img_dir"]), wb=out_wb)

    if save:
        out_wb.save(_output_path(paths))
    log.info("Done: post-processing completed in %s.", _elapsed_since(started_at))
    return results


def _collect_indicator_tables(
    results: dict[str, Any],
    inputs: dict[str, Any],
    extra_tables: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    """Merge the consolidated report tables (``extra_tables``, from
    ``write_results``) with the raw per-need indicator tables from
    ``build_fna_indicators`` and the DSO/TSO/Article-14 network tables, so
    CLI commands (``compute-fna-indicators``, ``fine-tune``, ``make-report``)
    have a single dict of named DataFrames to export/report on."""

    tables: dict[str, pd.DataFrame] = dict(extra_tables)
    frames = inputs["frames"]
    control = inputs.get("control") or {}

    try:
        from fna.io.indicators.core import build_fna_indicators

        tables.update(build_fna_indicators(
            results["residual"], frames.get("hours"), frames.get("days"),
            market_time_unit_minutes=float(control.get("market_time_unit_minutes", 60) or 60),
            flex_availability=frames.get("flex_profiles"), flex=frames.get("flex"),
            storage=results.get("storage"), prequalification=frames.get("prequalification"),
            target_year=inputs.get("target_year"),
        ))
    except Exception as exc:
        log.warning("Could not build per-need FNA indicator tables: %s", exc)

    try:
        from fna.io.indicators.network import compute_dso_needs, compute_fine_tuning_needs, compute_tso_needs

        tables["dso_needs"] = compute_dso_needs(results["residual"], frames.get("hours"), frames.get("dso_zones"))
        network_csv = results.get("network")
        if isinstance(network_csv, pd.DataFrame) and not network_csv.empty:
            tables["tso_needs"] = compute_tso_needs(network_csv, frames.get("network"), frames.get("hours"))
            tables["fine_tuning"] = compute_fine_tuning_needs(network_csv, frames.get("network"), frames.get("hours"))
    except Exception as exc:
        log.warning("Could not build DSO/TSO/Article-14 network-need tables: %s", exc)

    return tables


def run_monte_carlo(
    wb: Any | None = None,
    mc_params: dict[str, Any] | None = None,
    started_at: float | None = None,
    inputs_base: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
    manifest: Any | None = None,
) -> dict[int, dict]:
    """Run the full Monte Carlo workflow.

    ``inputs_base``, ``paths`` and ``sheet_suffix`` mirror run_optimisation:
    they let multi_year.py run the MC workflow once per target year, writing
    scenarios to a per-year folder and results to suffixed Excel sheets.
    Returns the per-scenario MC results dict.

    When ``paths`` is None an isolated run folder is minted
    (``data/outputs/runs/<run_id>/``) with per-scenario sub-folders and a
    provenance/compute manifest (including per-scenario solve timings).
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 Monte Carlo run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    mc_params = dict(mc_params or read_uncertainty_params(wb))
    status_cell = _safe_status_cell(wb)

    def status(message: str) -> None:
        _set_status(status_cell, message)
        log.info(message)

    owns_run = paths is None
    out_wb = None
    try:
        n_scenarios = int(mc_params["n_mc_scenarios"])
        max_workers = int(mc_params["max_parallel_workers"])
        seed = mc_params.get("seed_random")
        use_pecd = bool(mc_params.get("use_pecd_data"))

        if inputs_base is None:
            status("Reading base case inputs...")
            inputs_base = read_inputs_with_mc(wb, scenario_id=None)

        if owns_run:
            paths, manifest = begin_run("monte_carlo", inputs_base, mc_params=mc_params, n_scenarios=n_scenarios)
        paths["gms_file"] = resolve_gms_path(inputs_base.get("control"))
        out_wb = _open_output_workbook(out_path=_output_path(paths))
        _ensure_directories(paths)

        mc_params["pecd_target_year"] = int(mc_params.get("pecd_target_year") or inputs_base["target_year"])
        wind_portfolios = _wind_portfolios(inputs_base)
        _validate_wind_capacity(mc_params.get("wind_capacity_mw"), _wind_capacity_mw(inputs_base, wind_portfolios))

        status("Generating PECD load, wind and solar uncertainty scenarios...")
        mc_gen = MonteCarloScenarios(n_scenarios=n_scenarios, seed=seed)
        existing_scenarios = None
        rep_hours = _representative_hour_map(inputs_base)
        demand_pecd, solar_pecd, common_years = _load_pecd_load_solar(mc_params, inputs_base)
        if use_pecd:
            existing_scenarios = _generate_pecd_scenarios(mc_gen, mc_params, rep_hours, wind_portfolios, common_years)
        else:
            log.info("use_pecd_data is false; wind uses workbook mean/base profile.")

        uncertainty_scenarios = _add_pecd_load_solar_scenarios(
            mc_gen,
            inputs_base,
            rep_hours,
            demand_pecd,
            solar_pecd,
            existing_scenarios=existing_scenarios,
        )

        status(f"Running {n_scenarios} scenarios...")
        mc_results = _run_scenarios(
            n_scenarios,
            max_workers,
            inputs_base,
            uncertainty_scenarios,
            paths=paths,
            manifest=manifest,
        )
        if not mc_results:
            raise RuntimeError("All scenarios failed.")

        status("Writing Monte Carlo results...")
        write_mc_results_to_excel(out_wb, mc_results, requested_scenarios=n_scenarios, sheet_suffix=sheet_suffix)

        status("Generating charts...")
        chart_paths = _generate_charts(mc_results, uncertainty_scenarios, n_scenarios, max_workers, img_dir=Path(paths["img_dir"]))
        write_mc_charts_to_excel(out_wb, chart_paths, len(mc_results), n_scenarios, sheet_suffix=sheet_suffix)

        if owns_run:
            finish_run(manifest, paths, out_wb, status="completed")
        if save:
            out_wb.save(_output_path(paths))
        status(f"Done: Monte Carlo completed with {len(mc_results)} successful scenarios in {_elapsed_since(started_at)}")
        log.info("Base case cost: %.0f EUR", _get_base_cost(mc_results))
        log.info("Cost std dev: %.0f EUR", _get_cost_std(mc_results))
        log.info("Results written to %s", _output_path(paths))
        if owns_run:
            log.info("Run outputs in %s", paths["run_dir"])
        return mc_results
    except Exception as exc:
        status("Error: see the run's logs / gams_run.log")
        log.exception("Monte Carlo failed: %s", exc)
        if owns_run:
            finish_run(manifest, paths, None, status="failed", error=str(exc))
        raise


def run_single_scenario(
    scenario_id: int,
    inputs_base: dict,
    uncertainty_scenarios: dict[int, pd.DataFrame],
    paths: dict[str, str],
    inc_dir_scenario: Path,
    out_dir_scenario: Path,
) -> dict:
    """Run one scenario in a worker process without opening Excel. The return
    dict carries compute timing (``started_at``/``ended_at``/``wall_seconds`` and
    the GAMS solve time) so the parent can record it on the run manifest."""

    from fna.run_metadata import _iso, _now

    t0 = time.perf_counter()
    started_dt = _now()
    try:
        log.info("[Scenario %s] Starting", scenario_id)
        inputs = apply_uncertainty_scenario(inputs_base, scenario_id, uncertainty_scenarios)

        scenario_paths = dict(paths)
        scenario_paths["inc_dir"] = str(inc_dir_scenario)
        scenario_paths["out_dir"] = str(out_dir_scenario)
        scenario_paths["log_file"] = str(out_dir_scenario / "gams_run.log")
        results = run_model(inputs, scenario_paths)

        cost = _result_indicator(results, "total_cost", np.nan)
        timing = results.get("_gams_timing") or {}
        log.info("[Scenario %s] Complete (cost: %.0f EUR)", scenario_id, cost)
        return {
            "scenario_id": scenario_id, "results": results, "cost": cost, "error": None,
            "started_at": _iso(started_dt), "ended_at": _iso(_now()),
            "wall_seconds": round(time.perf_counter() - t0, 3),
            "gams_wall_seconds": timing.get("gams_wall_seconds"),
            "gams_resource_seconds": timing.get("gams_resource_seconds"),
        }
    except Exception as exc:
        log.error("[Scenario %s] Failed: %s", scenario_id, exc, exc_info=True)
        return {
            "scenario_id": scenario_id, "results": None, "cost": np.nan, "error": str(exc),
            "started_at": _iso(started_dt), "ended_at": _iso(_now()),
            "wall_seconds": round(time.perf_counter() - t0, 3),
            "gams_wall_seconds": None, "gams_resource_seconds": None,
        }


def _use_monte_carlo(params: dict[str, Any]) -> bool:
    return bool(params.get("run_monte_carlo", False))


def _configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, mode="w", encoding="utf-8"), logging.StreamHandler(sys.stdout)],
        force=True,
    )


def _open_workbook():
    """Load the input workbook headlessly with openpyxl (read-only).

    No live Excel/COM is required, so the model runs on any platform including
    headless Linux containers. ``data_only=True`` reads cached cell values; the
    FNA workbooks contain plain values (no live formulas) so this is exact.
    """

    wb_path = resolve_input_workbook(EXCEL_FILENAME)
    return load_workbook(wb_path, read_only=True, data_only=True)


def _open_output_workbook(input_wb: Any = None, out_path: Path | None = None):
    """Open (or create) the ``<input stem>-output.xlsx`` workbook used for all
    results/charts. Loads the existing file when present (so post-process runs
    update it in place); otherwise starts a clean, empty workbook.

    ``out_path`` selects the workbook location; when omitted it falls back to the
    legacy ``data/outputs/<stem>-output.xlsx`` (used by the standalone fna_* helpers).
    Run-isolated calls pass the per-run path so each run writes its own workbook.
    """

    out_path = Path(out_path) if out_path is not None else output_excel_path()
    if out_path.exists():
        return load_workbook(out_path)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty "Sheet"; real sheets are added on write
    return wb


def _ensure_directories(paths: dict[str, Any] | None = None) -> None:
    paths = paths or PATHS
    for path in [paths["out_dir"], paths["img_dir"], paths["log_file"].parent, paths["inc_dir"]]:
        Path(path).mkdir(parents=True, exist_ok=True)


def _reset_output_workbook(input_wb: Any = None) -> None:
    """Delete the previous ``<input stem>-output.xlsx`` so each successful run
    starts from a clean workbook."""

    out_path = output_excel_path()
    if out_path.exists():
        try:
            out_path.unlink()
        except OSError as exc:
            log.warning("Could not remove previous output workbook %s: %s", out_path, exc)


def _clean_generated_outputs(paths: dict[str, Any] | None = None) -> None:
    """Delete generated scenarios, images and old logs before each run, so a
    new run never reuses a previous run's results."""

    paths = paths or PATHS
    for root in [Path(paths["inc_dir"]), Path(paths["out_dir"])]:
        if not root.exists():
            continue
        for path in root.glob("scenario_*"):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
        scenarios_root = root / "scenarios"
        if scenarios_root.exists():
            shutil.rmtree(scenarios_root, ignore_errors=True)

    img_dir = Path(paths["img_dir"])
    if img_dir.exists():
        for image in img_dir.glob("*.png"):
            try:
                image.unlink()
            except OSError:
                pass

    csv_dir = csv_output_dir(Path(paths["out_dir"]))
    if csv_dir.exists():
        for csv_path in csv_dir.glob("*.csv"):
            try:
                csv_path.unlink()
            except OSError:
                pass

    logs_dir = Path(paths["log_file"]).parent
    if logs_dir.exists():
        for pattern in ["*.log", "*.lst"]:
            for log_path in logs_dir.glob(pattern):
                try:
                    log_path.unlink()
                except OSError:
                    pass


def _generate_pecd_scenarios(
    mc_gen: MonteCarloScenarios,
    mc_params: dict,
    rep_hours: list[RepresentativeHourMap],
    wind_portfolios: pd.DataFrame,
    common_years: list[int],
) -> dict[int, pd.DataFrame]:
    pecd_dir = _resolve_pecd_dir(mc_params)
    reader = PECDReader(pecd_dir)
    target_year = int(mc_params["pecd_target_year"])

    frames = []
    for wind_type in sorted(wind_portfolios["wind_type"].unique()):
        frames.append(reader.read_wind_time_series(PECD_COUNTRY_CODE, wind_type, target_year, weather_years=common_years))

    pecd_cf = pd.concat(frames, ignore_index=True)
    return mc_gen.generate_from_pecd(pecd_cf, rep_hours, wind_portfolios)


def _load_pecd_load_solar(
    mc_params: dict,
    inputs_base: dict,
) -> tuple[pd.DataFrame, pd.DataFrame, list[int]]:
    pecd_dir = _resolve_pecd_dir(mc_params)
    reader = PECDReader(pecd_dir)
    requested_years = PECD_YEARS or None
    target_year = int(mc_params.get("pecd_target_year") or inputs_base["target_year"])
    demand_pecd = reader.read_demand_time_series(PECD_COUNTRY_CODE, target_year, weather_years=requested_years)
    solar_pecd = reader.read_solar_time_series(PECD_COUNTRY_CODE, target_year, weather_years=requested_years)
    demand_years = set(int(y) for y in demand_pecd["weather_year"].dropna().unique())
    solar_years = set(int(y) for y in solar_pecd["weather_year"].dropna().unique())
    common_years = sorted(demand_years & solar_years)
    if not common_years:
        raise ValueError("No common PECD weather years between demand and solar data.")
    log.info("Using %d common PECD demand/solar weather years: %s-%s", len(common_years), common_years[0], common_years[-1])
    return demand_pecd, solar_pecd, common_years


def _add_pecd_load_solar_scenarios(
    mc_gen: MonteCarloScenarios,
    inputs_base: dict,
    rep_hours: list[RepresentativeHourMap],
    demand_pecd: pd.DataFrame,
    solar_pecd: pd.DataFrame,
    existing_scenarios: dict[int, pd.DataFrame] | None,
) -> dict[int, pd.DataFrame]:
    return mc_gen.add_pecd_load_solar(
        inputs_base,
        rep_hours,
        demand_pecd,
        solar_pecd,
        existing_scenarios=existing_scenarios,
    )


def _project_path(path_value: object) -> Path:
    path = Path(str(path_value)).expanduser()
    if path.is_absolute():
        return path
    return PROJECT_ROOT / path


def _resolve_pecd_dir(mc_params: dict) -> Path:
    configured = mc_params.get("pecd_data_dir")
    path = _project_path(configured) if configured else Path(PECD_DATA_DIR)
    if path.exists():
        return path
    default = Path(PECD_DATA_DIR)
    if default.exists():
        log.warning("Configured PECD data directory %s does not exist; using %s", path, default)
        return default
    return path


def _run_scenarios(
    n_scenarios: int,
    max_workers: int,
    inputs_base: dict,
    uncertainty_scenarios: dict[int, pd.DataFrame],
    paths: dict[str, Any] | None = None,
    manifest: Any | None = None,
) -> dict[int, dict]:
    paths = paths or PATHS
    scenario_dirs = {}
    scenarios_root = Path(paths.get("run_dir") or paths["out_dir"]) / "scenarios"
    for scenario_id in range(n_scenarios):
        out_dir = scenarios_root / f"scenario_{scenario_id}"
        inc_dir = out_dir / "inc"
        scenario_dirs[scenario_id] = (inc_dir, out_dir)

    paths_dict = {key: str(value) for key, value in paths.items()}
    mc_results: dict[int, dict] = {}

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for scenario_id in range(n_scenarios):
            inc_dir, out_dir = scenario_dirs[scenario_id]
            future = executor.submit(
                run_single_scenario,
                scenario_id,
                inputs_base,
                uncertainty_scenarios,
                paths_dict,
                inc_dir,
                out_dir,
            )
            futures[future] = scenario_id

        completed = 0
        for future in as_completed(futures):
            result = future.result()
            scenario_id = result["scenario_id"]
            _record_scenario_timing(manifest, result)
            if result["error"] is None:
                mc_results[scenario_id] = result["results"]
                completed += 1
                log.info("Scenarios completed: %d/%d", completed, n_scenarios)
            else:
                log.error("Scenario %s failed: %s", scenario_id, result["error"])

    log.info("Completed %d/%d scenarios successfully", len(mc_results), n_scenarios)
    return mc_results


def _record_scenario_timing(manifest: Any, result: dict[str, Any]) -> None:
    """Append one scenario's compute timing to the run manifest."""
    if manifest is None:
        return
    try:
        from fna.run_metadata import ScenarioTiming

        manifest.add_scenario(ScenarioTiming(
            scenario_id=int(result["scenario_id"]),
            started_at=result.get("started_at"),
            ended_at=result.get("ended_at"),
            wall_seconds=result.get("wall_seconds"),
            gams_wall_seconds=result.get("gams_wall_seconds"),
            gams_resource_seconds=result.get("gams_resource_seconds"),
            status="ok" if result.get("error") is None else "failed",
            error=result.get("error"),
        ))
    except Exception as exc:
        log.debug("Could not record scenario timing: %s", exc)


def _generate_charts(
    mc_results: dict[int, dict],
    uncertainty_scenarios: dict[int, pd.DataFrame],
    n_scenarios: int,
    max_workers: int,
    img_dir: Path | None = None,
) -> dict[str, Path]:
    return generate_mc_summary_charts(mc_results, uncertainty_scenarios, n_scenarios, max_workers, img_dir or PATHS["img_dir"])


def _representative_hour_map(inputs: dict) -> list[RepresentativeHourMap]:
    hours = inputs["frames"]["hours"].copy()
    days = inputs["frames"]["days"].copy()
    required_hours = {"time_id", "rep_day_id", "hour"}
    required_days = {"rep_day_id", "description"}
    if not required_hours.issubset(hours.columns):
        raise KeyError(f"02_RepHours missing columns: {sorted(required_hours - set(hours.columns))}")
    if not required_days.issubset(days.columns):
        raise KeyError(f"03_RepDays missing columns: {sorted(required_days - set(days.columns))}")

    rep_dates = {}
    for row in days.itertuples(index=False):
        date = _parse_representative_date(getattr(row, "description"))
        if date is not None:
            rep_dates[str(getattr(row, "rep_day_id"))] = date

    if not rep_dates:
        raise ValueError("Could not parse representative dates from 03_RepDays.description. Run rep_days.py first.")

    mapping = []
    for row in hours.itertuples(index=False):
        rep_day_id = str(getattr(row, "rep_day_id"))
        if rep_day_id not in rep_dates:
            raise ValueError(f"No representative date found for {rep_day_id}")
        date = rep_dates[rep_day_id]
        mapping.append(
            RepresentativeHourMap(
                time_id=str(getattr(row, "time_id")),
                month=int(date.month),
                day=int(date.day),
                hour=int(float(getattr(row, "hour"))),
            )
        )
    return mapping


def _parse_representative_date(text: object) -> pd.Timestamp | None:
    match = pd.Series([str(text)]).str.extract(r"((?:19|20)\d{2}-\d{2}-\d{2})").iloc[0, 0]
    if pd.isna(match):
        return None
    return pd.Timestamp(match)


def _wind_portfolios(inputs: dict) -> pd.DataFrame:
    res = inputs["frames"]["res"].copy()
    required = {"res_id", "technology"}
    if not required.issubset(res.columns):
        raise KeyError(f"07_RES_Portfolios missing columns: {sorted(required - set(res.columns))}")

    if WIND_RESOURCE_IDS:
        wind = res[res["res_id"].astype(str).isin(WIND_RESOURCE_IDS)].copy()
    else:
        wind = res[res["technology"].astype(str).str.lower().str.contains("wind", na=False)].copy()

    if wind.empty:
        raise ValueError("No wind portfolios found in 07_RES_Portfolios.")

    wind["wind_type"] = wind["technology"].map(_wind_type_from_technology)
    if wind["wind_type"].isna().any():
        bad = wind.loc[wind["wind_type"].isna(), ["res_id", "technology"]].to_dict("records")
        raise ValueError(f"Cannot classify wind portfolios as onshore/offshore: {bad}")
    return wind[["res_id", "technology", "wind_type"]].copy()


def _wind_type_from_technology(value: object) -> str | None:
    text = str(value).lower()
    if "offshore" in text:
        return "offshore"
    if "onshore" in text or "wind" in text:
        return "onshore"
    return None


def _wind_capacity_mw(inputs: dict, wind_portfolios: pd.DataFrame) -> float:
    target_year = int(inputs["target_year"])
    capacity_col = f"capacity_mw_{target_year}"
    res = inputs["frames"]["res"].copy()
    if capacity_col not in res.columns:
        raise KeyError(f"Missing {capacity_col} in 07_RES_Portfolios.")

    wind_ids = set(wind_portfolios["res_id"].astype(str))
    capacity = pd.to_numeric(
        res.loc[res["res_id"].astype(str).isin(wind_ids), capacity_col],
        errors="coerce",
    ).fillna(0.0).sum()
    if capacity <= 0:
        raise ValueError(f"Derived wind capacity is {capacity}; check 07_RES_Portfolios.")
    return float(capacity)


def _validate_wind_capacity(configured_capacity: object, derived_capacity: float) -> None:
    if configured_capacity is None or pd.isna(configured_capacity):
        log.info("Derived wind capacity from 07_RES_Portfolios: %.2f MW", derived_capacity)
        return

    configured = float(configured_capacity)
    if configured <= 0:
        raise ValueError("wind_capacity_mw must be positive if supplied.")

    rel_diff = abs(configured - derived_capacity) / derived_capacity
    if rel_diff > 0.01:
        raise ValueError(
            f"wind_capacity_mw={configured:.2f} does not match 07_RES_Portfolios "
            f"wind capacity={derived_capacity:.2f} MW."
        )
    log.info("Validated configured wind capacity %.2f MW", configured)


def _safe_status_cell(wb):
    # Headless runs have no live Excel to surface progress in; status is logged
    # instead (see the status() closures), so there is no status cell to write.
    return None


def _set_status(cell: Any, message: str) -> None:
    # No-op: progress is reported through logging in headless mode.
    return None


def _elapsed_since(started_at: float | None) -> str:
    if started_at is None:
        return "unknown time"
    return _format_elapsed(time.perf_counter() - started_at)


def _format_elapsed(seconds: float) -> str:
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def _get_base_cost(mc_results: dict[int, dict]) -> float:
    if 0 in mc_results:
        return _result_indicator(mc_results[0], "total_cost", np.nan)
    return np.nan


def _get_cost_std(mc_results: dict[int, dict]) -> float:
    costs = []
    for result in mc_results.values():
        cost = _result_indicator(result, "total_cost")
        if cost is not None and not pd.isna(cost):
            costs.append(float(cost))
    return float(np.std(costs)) if len(costs) > 1 else 0.0


def _result_indicator(result: dict, metric: str, default: object = None) -> object:
    indicators = result.get("indicators", {})
    if isinstance(indicators, dict):
        return indicators.get(metric, default)
    if isinstance(indicators, pd.DataFrame) and {"metric", "value"}.issubset(indicators.columns):
        rows = indicators[indicators["metric"].astype(str).str.strip().eq(metric)]
        if not rows.empty:
            return pd.to_numeric(rows.iloc[0]["value"], errors="coerce")
    return default


if __name__ == "__main__":
    main()
