"""
styling.py - flag input data that needs user attention.

When a workbook is generated for a country/year, the hourly time-series come
from ENTSO-E (historical), but the *structural* inputs - thermal-fleet technical
limits (Pmin / ramp / start-up), flexibility & storage portfolios, interconnector
definitions, network needs - are scenario **assumptions** that must be reviewed
and confirmed for the chosen country/scenario.

``highlight_assumptions`` colour-codes the workbook so those are easy to spot:

* sheet tabs   - amber if the sheet holds structural assumptions to confirm,
                 green if every row is historical/empirical, else untouched;
* data rows    - amber fill where the row's ``data_quality`` marks it an
                 assumption/placeholder/proxy (vs. historical/actual/empirical).

It returns a per-sheet summary so the CLI can print a "confirm these" notice.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

# Sheets whose values are scenario assumptions a user should confirm per country.
ATTENTION_SHEETS = {
    "06_DispatchableBlocks": "thermal fleet technical limits (Pmin, ramp, min up/down, start-up cost)",
    "09_FlexStorage": "flexibility / storage portfolios (power, energy, efficiency, cost)",
    "04_Interconnectors": "interconnector capacities and definitions",
    "07_RES_Portfolios": "RES installed capacities and curtailable shares",
    "13_NetworkNeeds": "Article-14 / network needs",
    "13b_DSO_Zones": "DSO hosting-capacity zones",
    "10b_Prequalification_Log": "flexibility prequalification status",
    "17_Barriers_Digitalisation": "barriers & digitalisation register",
}

# data_quality substrings that mean "assumed" (amber) vs. "measured" (ok).
_ASSUMPTION_KEYS = ("assumption", "placeholder", "proxy", "model design", "assumed", "estimate")
_HISTORICAL_KEYS = ("historical", "actual", "empirical", "measured", "entso")

AMBER = PatternFill("solid", fgColor="FFE699")   # row fill: needs attention
AMBER_TAB = "FFC000"                              # sheet tab: assumptions to confirm
GREEN_TAB = "70AD47"                              # sheet tab: fully historical


def _is_assumption(value: Any) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    if any(k in text for k in _HISTORICAL_KEYS):
        return False
    return any(k in text for k in _ASSUMPTION_KEYS) or text not in ("", "n/a")


def highlight_assumptions(workbook: "openpyxl.Workbook | str | Path") -> list[dict[str, Any]]:
    """Colour-code assumption sheets/rows in ``workbook`` (path or open Workbook).

    Saves in place when given a path. Returns a list of
    ``{sheet, why, n_assumption_rows, n_total_rows}`` for the sheets that need
    attention, so the caller can print a confirmation notice."""

    path = None
    if isinstance(workbook, (str, Path)):
        path = Path(workbook)
        wb = openpyxl.load_workbook(path)
    else:
        wb = workbook

    summary: list[dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1), [])]
        dq_col = header.index("data_quality") + 1 if "data_quality" in header else None

        n_assumption = 0
        n_total = 0
        if dq_col is not None:
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=dq_col).value
                if val is None or str(val).strip() == "":
                    continue
                n_total += 1
                if _is_assumption(val):
                    n_assumption += 1
                    for col in range(1, len(header) + 1):
                        ws.cell(row=row, column=col).fill = AMBER

        if name in ATTENTION_SHEETS:
            ws.sheet_properties.tabColor = AMBER_TAB
            summary.append({
                "sheet": name,
                "why": ATTENTION_SHEETS[name],
                "n_assumption_rows": n_assumption,
                "n_total_rows": n_total,
            })
        elif dq_col is not None and n_total > 0 and n_assumption == 0:
            ws.sheet_properties.tabColor = GREEN_TAB
        elif n_assumption > 0:
            ws.sheet_properties.tabColor = AMBER_TAB

    if path is not None:
        wb.save(path)
    return summary


def print_attention_notice(summary: list[dict[str, Any]], echo=print) -> None:
    """Print a human-readable 'confirm these inputs' notice from the summary."""
    if not summary:
        return
    echo("\n  Review these sheets before running (amber tabs / rows are assumptions to confirm):")
    for item in summary:
        extra = f" - {item['n_assumption_rows']}/{item['n_total_rows']} rows flagged" if item["n_total_rows"] else ""
        echo(f"    - {item['sheet']}: {item['why']}{extra}")
    echo("  Hourly time-series (demand/wind/solar/flows/prices) are historical ENTSO-E and need no review.\n")
