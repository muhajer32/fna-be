"""
full_year.py - build a full-year hourly FNA input workbook from cached ENTSO-E
data for a user-selected country code and data year.

This produces ``data/inputs/excel/<CC>_FullYear<year>.xlsx`` with the same
sheet/column schema as the template workbook, but:

  - 02_RepHours / 03_RepDays: one row-set per calendar day (D001..D365/366),
    weight_days=1 each, instead of 10 clustered representative days. This is
    "no clustering" while staying inside the existing rep_day_id/weight_days
    contract (see io_excel.write_inc_files), so the GAMS model can run on it
    unmodified -- with the caveat noted in 00_ReadMe that storage SOC cycles
    are currently enforced per calendar day (same_day_pairs), not per year.
  - 08_RES_CF_Profiles, 05_IntercoProfiles, 10_FlexAvailability,
    11_Availability_Outages, 12_Reserve_ForecastError: re-keyed to the same
    full-year time_ids, populated with real ENTSO-E series where available
    (demand, wind/solar CF, cross-border flows, day-ahead prices, forecast
    errors, generation outages) and documented robust assumptions elsewhere
    (flexible-resource availability, reserve dimensioning) per
    docs/METHODOLOGY.md (§6).
  - 04/06/07/09/13/13b/17/20/01_Control/00_ReadMe: lightly refreshed
    source_id/data_quality/notes; numeric assumptions only changed where a
    real ENTSO-E figure is a clear improvement (documented per row).

Run after ``fna fetch --country <CC> --year <YYYY>`` has populated
``data/inputs/raw_<CC><YYYY>/``, or use ``build-full-year --fetch`` to pull
missing cached data automatically.
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


from fna.config import ENTSOE_TIMEZONE, EXCEL_DIR, EXCEL_FILENAME, INPUTS_DIR, PROJECT_ROOT, raw_data_dir, resolve_input_workbook  # noqa: E402
from fna.inputs.styling import highlight_assumptions, print_attention_notice  # noqa: E402

COUNTRY_CODE = "BE"
DATA_YEAR = 2023
BUILD_YEAR = 2023

RAW = raw_data_dir("BE", 2023)
SRC_WB = resolve_input_workbook(EXCEL_FILENAME)
OUT_WB = EXCEL_DIR / "BE_FullYear2023.xlsx"

N_DAYS = 365
N_HOURS = N_DAYS * 24  # 8760
INDEX = pd.date_range("2023-01-01", periods=N_HOURS, freq="h", tz="UTC")
CALENDAR_INDEX = INDEX.tz_convert(ENTSOE_TIMEZONE or "Europe/Brussels")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

INSTALLED = {
    "nuclear": 4937.0,
    "fossil_gas": 6987.62,
    "biomass": 730.34,
    "waste": 383.64,
    "wind_offshore": 2262.1,
    "wind_onshore": 3053.18,
    "solar": 6474.96,
    "hydro_psp": 1308.0,
    "hydro_ror": 185.51,
}
_DEFAULT_INSTALLED = dict(INSTALLED)


def _configure(country: str, year: int, template: "Path | None", data_year: "int | None") -> str:
    """Set the module globals for one (country, year) build and return the
    output-workbook path's country code. The hourly cache lives under
    ``data/inputs/raw_<cc><data_year>/`` and the output is ``<CC>_FullYear<year>.xlsx``;
    structural sheets + installed wind/solar capacities come from ``template``."""

    global BUILD_YEAR, CALENDAR_INDEX, COUNTRY_CODE, DATA_YEAR, RAW, SRC_WB, OUT_WB, INDEX, N_DAYS, N_HOURS, INSTALLED
    cc = str(country).strip().upper()
    dy = int(data_year or year)
    COUNTRY_CODE = cc
    DATA_YEAR = dy
    BUILD_YEAR = int(year)

    # Hourly cache dir: prefer current raw_<CC><year>, but accept older spellings.
    candidates = [
        raw_data_dir(cc, dy),
        INPUTS_DIR / f"raw_{cc.lower()}{dy}",
        INPUTS_DIR / f"raw_{cc}_{dy}",
        INPUTS_DIR / f"raw_{cc.lower()}_{dy}",
    ]
    RAW = next((c for c in candidates if c.exists()), candidates[0])

    SRC_WB = Path(template) if template else resolve_input_workbook(EXCEL_FILENAME)
    OUT_WB = EXCEL_DIR / f"{cc}_FullYear{year}.xlsx"
    local_tz = ENTSOE_TIMEZONE or "Europe/Brussels"
    CALENDAR_INDEX = pd.date_range(f"{dy}-01-01", f"{dy + 1}-01-01", freq="h", tz=local_tz, inclusive="left")
    INDEX = CALENDAR_INDEX.tz_convert("UTC")
    N_HOURS = len(INDEX)
    N_DAYS = N_HOURS // 24

    INSTALLED = dict(_DEFAULT_INSTALLED)
    INSTALLED.update(_installed_from_template(SRC_WB))
    return cc


def _installed_from_template(template: "Path") -> dict[str, float]:
    """Read wind/solar installed capacities from the template's 07_RES_Portfolios
    so capacity factors (= actual / installed) are country-correct. Thermal/hydro
    capacities keep the documented defaults (flagged for review)."""

    try:
        res = pd.read_excel(template, sheet_name="07_RES_Portfolios")
    except Exception:
        return {}
    cap_col = next((c for c in res.columns if str(c).startswith("capacity_MW")), None)
    if not cap_col or "technology" not in res.columns:
        return {}
    tech = res["technology"].astype(str).str.lower()
    cap = pd.to_numeric(res[cap_col], errors="coerce").fillna(0.0)
    out: dict[str, float] = {}
    for key, needle in [("wind_offshore", "offshore"), ("wind_onshore", "onshore"), ("solar", "solar")]:
        total = float(cap[tech.str.contains(needle, na=False)].sum())
        if total > 0:
            out[key] = total
    return out


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_aligned(name: str, required: bool = True) -> pd.DataFrame:
    path = RAW / f"{name}.csv"
    if not path.exists():
        if required:
            raise FileNotFoundError(f"Required cached ENTSO-E series not found: {path}")
        return pd.DataFrame(index=INDEX)
    df = pd.read_csv(path, index_col=0)
    df.index = pd.to_datetime(df.index, utc=True)
    df = df[~df.index.duplicated(keep="first")]
    df = df.reindex(INDEX)
    df = df.ffill().bfill()
    return df


def season_from_month(month: int) -> str:
    if month in (12, 1, 2):
        return "winter"
    if month in (3, 4, 5):
        return "spring"
    if month in (6, 7, 8):
        return "summer"
    return "autumn"


def day_type_from_date(ts: pd.Timestamp) -> str:
    return "weekday" if ts.weekday() < 5 else "weekend"


def build_calendar() -> pd.DataFrame:
    cal = pd.DataFrame(index=INDEX)
    day_no = (np.arange(N_HOURS) // 24) + 1
    hour = np.arange(N_HOURS) % 24
    cal["day_no"] = day_no
    cal["hour"] = hour
    cal["rep_day_id"] = [f"D{d:03d}" for d in day_no]
    cal["time_id"] = [f"D{d:03d}_H{h:02d}" for d, h in zip(day_no, hour)]
    cal["next_time_id"] = [
        f"D{d:03d}_H{(h + 1) % 24:02d}" for d, h in zip(day_no, hour)
    ]
    cal["season"] = [season_from_month(ts.month) for ts in CALENDAR_INDEX]
    cal["day_type"] = [day_type_from_date(ts) for ts in CALENDAR_INDEX]
    cal["date"] = CALENDAR_INDEX.date
    return cal


def umm_availability(plant_type: str, capacity_mw: float) -> pd.Series:
    """Hourly availability_pct from real ENTSO-E UMM outage events."""

    df = pd.read_csv(RAW / "generation_unavailability.csv")
    sub = df[(df["plant_type"] == plant_type) & (df["docstatus"].isna())].copy()
    unavail = pd.Series(0.0, index=INDEX)
    if not sub.empty:
        sub["start"] = pd.to_datetime(sub["start"], utc=True)
        sub["end"] = pd.to_datetime(sub["end"], utc=True)
        sub["unavail_mw"] = (sub["nominal_power"] - sub["avail_qty"]).clip(lower=0.0)
        for _, row in sub.iterrows():
            mask = (INDEX >= row["start"]) & (INDEX < row["end"])
            unavail.loc[mask] += row["unavail_mw"]
    avail_pct = 1.0 - (unavail / capacity_mw).clip(lower=0.0, upper=0.95)
    return avail_pct


def header_style(ws) -> None:
    n_cols = ws.max_column
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def sheet_sort_key(name: str) -> tuple[int, str]:
    import re

    match = re.match(r"^(\d+)", str(name))
    if match:
        return int(match.group(1)), str(name).lower()
    return 999, str(name).lower()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_hours_and_days(cal: pd.DataFrame, load: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    data_label = f"ENTSO-E {DATA_YEAR} actual"
    hours = pd.DataFrame({
        "time_id": cal["time_id"],
        "rep_day_id": cal["rep_day_id"],
        "hour": cal["hour"],
        "next_time_id": cal["next_time_id"],
        "season": cal["season"],
        "day_type": cal["day_type"],
        "weight_days": 1,
        "weight_hours": 1,
        "chronology_group": cal["rep_day_id"],
        "gross_demand_MW_2025": load["load_actual"].to_numpy(),
        "source_id": "S25",
        "data_quality": f"historical ({data_label})",
        "notes": f"Full calendar year {DATA_YEAR}, no clustering; used as demand-level proxy.",
    })

    days_grp = cal.groupby("rep_day_id", sort=False).agg(
        date=("date", "first"), season=("season", "first"), day_type=("day_type", "first")
    ).reset_index()
    days = pd.DataFrame({
        "rep_day_id": days_grp["rep_day_id"],
        "description": [f"Calendar date {d}" for d in days_grp["date"]],
        "season": days_grp["season"],
        "day_type": days_grp["day_type"],
        "weight_days": 1,
        "selection_reason": "Actual calendar day - no clustering (full-year hourly dataset)",
        "probability_pct": 100.0 / N_DAYS,
        "source_id": "S25",
        "data_quality": f"historical ({data_label})",
    })
    return hours, days


def build_res_cf(cal: pd.DataFrame, res_actual: pd.DataFrame, res_portfolios: pd.DataFrame) -> pd.DataFrame:
    cf_wind_offshore = (res_actual["wind_offshore"] / INSTALLED["wind_offshore"]).clip(0, 1)
    cf_wind_onshore = (res_actual["wind_onshore"] / INSTALLED["wind_onshore"]).clip(0, 1)
    cf_solar = (res_actual["solar"] / INSTALLED["solar"]).clip(0, 1)

    cf_map = {
        "offshore_wind": (cf_wind_offshore, f"ENTSO-E {DATA_YEAR} actual generation / installed offshore wind capacity ({INSTALLED['wind_offshore']:.1f} MW)"),
        "onshore_wind": (cf_wind_onshore, f"ENTSO-E {DATA_YEAR} actual generation / installed onshore wind capacity ({INSTALLED['wind_onshore']:.1f} MW); single national profile applied to all onshore regions"),
        "solar_pv": (cf_solar, f"ENTSO-E {DATA_YEAR} actual generation / installed solar capacity ({INSTALLED['solar']:.1f} MW); single national profile applied to all PV segments"),
    }

    avail_offshore = umm_availability("Wind Offshore", INSTALLED["wind_offshore"])

    rows = []
    for _, res in res_portfolios.iterrows():
        tech = str(res["technology"]).strip().lower()
        if tech not in cf_map:
            continue
        cf_series, note = cf_map[tech]
        avail = avail_offshore if tech == "offshore_wind" else pd.Series(0.98, index=INDEX)
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "res_id": res["res_id"],
            "capacity_factor": cf_series.to_numpy(),
            "availability_pct": avail.to_numpy(),
            "source_id": "S26",
            "data_quality": f"historical (ENTSO-E {DATA_YEAR} actual)",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


def _neighbour_from_border_id(border_id: str) -> str | None:
    parts = [p for p in str(border_id).upper().replace("-", "_").split("_") if p]
    if COUNTRY_CODE in parts and len(parts) > 1:
        return next((p for p in parts if p != COUNTRY_CODE), None)
    return parts[-1] if parts else None


def _flow_col(flows: pd.DataFrame, border_id: str, direction: str) -> pd.Series:
    col = f"{border_id}_{direction}"
    if col in flows.columns:
        return flows[col]
    print(f"  warning: {col} missing from cross_border_flows.csv; using 0 MW fixed flow.")
    return pd.Series(0.0, index=INDEX)


def build_interco_profiles(cal: pd.DataFrame, flows: pd.DataFrame, prices: pd.DataFrame, nb_prices: pd.DataFrame, borders: pd.DataFrame) -> pd.DataFrame:
    local_price = prices["price_eur_mwh"]
    cap_by_border = borders.set_index("border_id")[["import_capacity_MW_2025", "export_capacity_MW_2025"]]

    rows = []
    for border_id in borders["border_id"]:
        exp = _flow_col(flows, str(border_id), "export")
        imp = _flow_col(flows, str(border_id), "import")
        net = imp - exp
        import_cap = float(cap_by_border.loc[border_id, "import_capacity_MW_2025"])
        export_cap = float(cap_by_border.loc[border_id, "export_capacity_MW_2025"])
        # Realised ENTSO-E physical flows occasionally exceed the
        # workbook's assumed NTC (loop-flow/PST effects); clip to
        # the assumed line capacity so fixed flows stay model-feasible.
        fixed_import = net.clip(lower=0.0, upper=import_cap)
        fixed_export = (-net).clip(lower=0.0, upper=export_cap)
        nb_code = _neighbour_from_border_id(str(border_id))
        nb_price = nb_prices.get(nb_code, local_price) if nb_code else local_price
        price_note = (
            f"{nb_code} day-ahead price" if nb_code and nb_code in nb_prices.columns
            else f"{COUNTRY_CODE} day-ahead price fallback (neighbour price unavailable)"
        )
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "border_id": border_id,
            "availability_pct": 0.95,
            "fixed_import_MW": fixed_import.to_numpy(),
            "fixed_export_MW": fixed_export.to_numpy(),
            "import_price_EUR_MWh": nb_price.to_numpy(),
            "export_price_EUR_MWh": local_price.to_numpy(),
            "source_id": "S27",
            "data_quality": f"historical (ENTSO-E {DATA_YEAR} actual flows/prices)",
            "notes": f"fixed_* = realised {DATA_YEAR} net flow, clipped to the import/export capacity assumption; import price uses {price_note}; availability_pct = tradable headroom assumption (model can still trade up to capacity x availability_pct beyond the fixed flow).",
        }))
    return pd.concat(rows, ignore_index=True)


def build_reserve_forecast_error(cal: pd.DataFrame, load: pd.DataFrame, res_actual: pd.DataFrame, res_fc: pd.DataFrame) -> pd.DataFrame:
    season = cal["season"].to_numpy()
    hour = cal["hour"].to_numpy()

    load_err = (load["load_actual"] - load["load_forecast"]) / load["load_actual"]
    wind_actual = res_actual["wind_onshore"] + res_actual["wind_offshore"]
    wind_fc = res_fc["wind_onshore_fc"] + res_fc["wind_offshore_fc"]
    wind_cap = INSTALLED["wind_onshore"] + INSTALLED["wind_offshore"]
    wind_err = (wind_actual - wind_fc) / wind_cap
    solar_err = (res_actual["solar"] - res_fc["solar_fc"]) / INSTALLED["solar"]

    df = pd.DataFrame({"season": season, "load_err": load_err.to_numpy(), "wind_err": wind_err.to_numpy(), "solar_err": solar_err.to_numpy()})
    season_stats = df.groupby("season").agg(
        load_pct=("load_err", lambda s: float(s.std())),
        wind_pct=("wind_err", lambda s: float(s.std())),
        solar_pct=("solar_err", lambda s: float(s.std())),
    )

    load_pct = np.array([season_stats.loc[s, "load_pct"] for s in season])
    wind_pct = np.array([season_stats.loc[s, "wind_pct"] for s in season])
    # Solar forecast error only matters during daylight hours; keep it ~0 overnight.
    daylight = (hour >= 7) & (hour <= 18)
    solar_pct = np.where(daylight, [season_stats.loc[s, "solar_pct"] for s in season], 0.005)

    # Reserve dimensioning (FCR/aFRR/mFRR): keep the existing template
    # seasonal/diurnal pattern (S20/S21), extended to all calendar days.
    # Winter evening peak (17-21h) gets the elevated band.
    winter_evening = (season == "winter") & (hour >= 17) & (hour <= 21)
    winter = season == "winter"

    fcr = np.full(N_HOURS, 90.0)
    afrr_up = np.where(winter_evening, 330.0, np.where(winter, 300.0, 250.0))
    afrr_dn = np.where(winter, 310.0, 250.0)
    mfrr_up = np.where(winter_evening, 1200.0, np.where(winter, 1000.0, 900.0))
    mfrr_dn = np.where(winter, 950.0, 750.0)
    largest_outage = np.where(winter_evening, 1000.0, 250.0)

    return pd.DataFrame({
        "time_id": cal["time_id"],
        "FCR_MW": fcr,
        "aFRR_up_MW": afrr_up,
        "aFRR_down_MW": afrr_dn,
        "mFRR_up_MW": mfrr_up,
        "mFRR_down_MW": mfrr_dn,
        "load_forecast_error_pct": load_pct,
        "wind_forecast_error_pct": wind_pct,
        "solar_forecast_error_pct": solar_pct,
        "largest_unit_outage_stress_MW": largest_outage,
        "source_id": "S20/S21/S28",
        "data_quality": f"load/wind/solar error = empirical std (ENTSO-E {DATA_YEAR} forecast vs actual, per season); reserve volumes = seasonal/diurnal assumption",
        "notes": "Forecast-error pct = std((actual-forecast)/reference) per season, reference=actual load resp. installed wind/solar capacity. Reserve MW pattern from the template methodology (S20/S21), now applied to every day of the year.",
    })


def build_availability_outages(cal: pd.DataFrame, dispatchable: pd.DataFrame) -> pd.DataFrame:
    nuclear_avail = umm_availability("Nuclear", INSTALLED["nuclear"])
    gas_avail = umm_availability("Fossil Gas", INSTALLED["fossil_gas"])
    biomass_avail = umm_availability("Biomass", INSTALLED["biomass"])

    season = cal["season"]
    # No UMM coverage for run-of-river hydro or industrial CHP: keep a mild,
    # documented seasonal assumption (lower hydro availability in low-flow
    # summer; CHP maintenance concentrated in shoulder seasons).
    hydro_avail = season.map({"winter": 0.97, "spring": 0.95, "summer": 0.85, "autumn": 0.95}).astype(float)
    chp_avail = season.map({"winter": 0.95, "spring": 0.88, "summer": 0.92, "autumn": 0.88}).astype(float)

    profiles = {
        "NUC_EXIST": (nuclear_avail, "S29", f"ENTSO-E {DATA_YEAR} UMM (REMIT) outage events for {COUNTRY_CODE} nuclear units"),
        "CCGT_EXIST": (gas_avail, "S29", f"ENTSO-E {DATA_YEAR} UMM outage events for {COUNTRY_CODE} Fossil Gas units"),
        "CCGT_NEW": (gas_avail, "S29", "Same UMM-derived profile as CCGT_EXIST (no separate unit-level split available)"),
        "OCGT_PEAKER": (gas_avail, "S29", "Same UMM-derived profile as CCGT_EXIST (no separate unit-level split available)"),
        "CHP_INDUSTRIAL": (chp_avail, "S30", "No UMM coverage for industrial CHP; seasonal maintenance assumption"),
        "BIOMASS_WASTE": (biomass_avail, "S29", f"ENTSO-E {DATA_YEAR} UMM outage events for {COUNTRY_CODE} Biomass units"),
        "HYDRO_RUN": (hydro_avail, "S30", "No UMM coverage for small hydro; seasonal low-flow assumption"),
    }

    rows = []
    for _, unit in dispatchable.iterrows():
        ugid = unit["unit_group_id"]
        if ugid not in profiles:
            continue
        avail, source_id, note = profiles[ugid]
        forced = (1.0 - avail).clip(lower=0.0)
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "unit_group_id": ugid,
            "planned_outage_pct": 0.0,
            "forced_outage_rate_pct": forced.to_numpy(),
            "availability_pct": avail.to_numpy(),
            "largest_unit_outage_MW": float(unit["capacity_per_block_MW_2025"]),
            "source_id": source_id,
            "data_quality": f"historical (ENTSO-E {DATA_YEAR} UMM)" if source_id == "S29" else "seasonal assumption",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


def build_flex_availability(cal: pd.DataFrame, flex: pd.DataFrame) -> pd.DataFrame:
    hour = cal["hour"].to_numpy()
    weekday = (cal["day_type"] == "weekday").to_numpy()
    season = cal["season"].to_numpy()
    month = np.array([ts.month for ts in CALENDAR_INDEX])

    psp_avail = umm_availability("Hydro Pumped Storage", INSTALLED["hydro_psp"])

    def diurnal(peak_hours, base, peak):
        return np.where(np.isin(hour, peak_hours), peak, base)

    # EV smart charging: cars plugged in mainly overnight + workplace midday;
    # discharge/V2G (up) availability lower than charge-shift (down).
    ev_down = diurnal(range(19, 24), 0.30, 0.65)
    ev_down = np.where(np.isin(hour, range(0, 7)), 0.70, ev_down)
    ev_up = diurnal(range(17, 22), 0.10, 0.30)
    ev_up = np.where(weekday, ev_up, ev_up * 0.7)

    # Heat-pump thermal shift: tracks heating season + early-morning pre-heat.
    hp_season_factor = pd.Series(season).map({"winter": 1.0, "autumn": 0.6, "spring": 0.6, "summer": 0.15}).to_numpy()
    hp_diurnal = diurnal(range(5, 9), 0.25, 0.45)
    hp_up = hp_diurnal * hp_season_factor
    hp_down = (hp_diurnal * 0.8) * hp_season_factor

    # Industrial DR: tied to production hours, reduced on weekends and during
    # a generic August industrial shutdown/holiday period.
    august = month == 8
    ind_fast = np.where(weekday & ~august, diurnal(range(6, 22), 0.65, 0.85), 0.10)
    ind_slow = np.where(weekday & ~august, 0.75, 0.15)

    # Commercial HVAC: business hours, weekdays only.
    hvac = np.where(weekday, diurnal(range(8, 19), 0.10, 0.55), 0.05)

    profiles = {
        "PSP_COO_PLATE_TAILLE": (psp_avail.to_numpy(), psp_avail.to_numpy(), 1.0, "S31", f"ENTSO-E {DATA_YEAR} UMM outage events for {COUNTRY_CODE} Hydro Pumped Storage"),
        "BESS_1H": (np.full(N_HOURS, 0.92), np.full(N_HOURS, 0.92), 1.0, "S32", "Technical availability assumption (battery fleet)"),
        "BESS_4H": (np.full(N_HOURS, 0.92), np.full(N_HOURS, 0.92), 1.0, "S32", "Technical availability assumption (battery fleet)"),
        "EV_SMART_CHARGING": (ev_up, ev_down, 1.0, "S33", "Diurnal plug-in pattern (overnight + workplace charging); weekday/weekend split"),
        "HEATPUMP_SHIFT": (hp_up, hp_down, 1.0, "S33", "Heating-season + early-morning pre-heat shift pattern"),
        "INDUSTRIAL_DR_FAST": (ind_fast, ind_fast, 1.0, "S34", "Weekday production-hour availability; reduced to standby during August shutdown/weekends"),
        "INDUSTRIAL_DR_SLOW": (ind_slow, ind_slow, 1.0, "S34", "Weekday production-hour availability; reduced to standby during August shutdown/weekends"),
        "COMMERCIAL_HVAC": (hvac, hvac, 1.0, "S34", "Weekday business-hours availability"),
    }

    rows = []
    for _, fx in flex.iterrows():
        fid = fx["flex_id"]
        if fid not in profiles:
            continue
        up, dn, budget, source_id, note = profiles[fid]
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "flex_id": fid,
            "availability_pct_up": up,
            "availability_pct_down": dn,
            "energy_budget_pct": budget,
            "source_id": source_id,
            "data_quality": f"historical (ENTSO-E {DATA_YEAR} UMM)" if source_id == "S31" else "robust assumption (documented diurnal/seasonal pattern)",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


# ---------------------------------------------------------------------------
# Static-sheet refreshes
# ---------------------------------------------------------------------------

def refresh_sources(sources: pd.DataFrame) -> pd.DataFrame:
    new_rows = [
        {"source_id": "S25", "source_name": f"ENTSO-E Transparency Platform - Total Load {DATA_YEAR}",
         "what_it_supports": f"02_RepHours full-year hourly demand (D001-D{N_DAYS:03d})",
         "url": "https://transparency.entsoe.eu/load-domain/r2/totalLoadR2/show", "notes": f"{COUNTRY_CODE} actual load, hourly mean of source values, calendar year {DATA_YEAR}"},
        {"source_id": "S26", "source_name": f"ENTSO-E Transparency Platform - Generation per type {DATA_YEAR}",
         "what_it_supports": "08_RES_CF_Profiles wind/solar capacity factors",
         "url": "https://transparency.entsoe.eu/generation/r2/actualGenerationPerProductionType/show", "notes": f"{COUNTRY_CODE} actual wind onshore/offshore/solar generation / installed capacity"},
        {"source_id": "S27", "source_name": f"ENTSO-E Transparency Platform - cross-border flows & day-ahead prices {DATA_YEAR}",
         "what_it_supports": "05_IntercoProfiles fixed flows and prices",
         "url": "https://transparency.entsoe.eu/transmission-domain/physicalFlow/show", "notes": f"Physical flows both directions per configured {COUNTRY_CODE} border; day-ahead prices use neighbour prices when cached, otherwise local price fallback"},
        {"source_id": "S28", "source_name": f"ENTSO-E day-ahead forecasts vs actuals {DATA_YEAR}",
         "what_it_supports": "12_Reserve_ForecastError empirical forecast-error std-dev",
         "url": "https://transparency.entsoe.eu/transmission-domain/r2/dayAheadAggregatedForecastsForGenerationAndLoad/show", "notes": "Per-season std of (actual-forecast)/reference for load, wind, solar; reserve MW volumes remain template assumptions (S20/S21)"},
        {"source_id": "S29", "source_name": f"ENTSO-E Transparency Platform - Unavailability of Generation Units (UMM/REMIT) {DATA_YEAR}",
         "what_it_supports": "11_Availability_Outages nuclear/gas/biomass availability profiles",
         "url": "https://transparency.entsoe.eu/outage-domain/r2/unavailabilityOfGenerationUnits/show", "notes": "Only REMIT-reporting units (>=100MW) are covered; OCGT/CHP/hydro-run units below threshold use the same fleet-level Fossil Gas profile or a seasonal assumption"},
        {"source_id": "S30", "source_name": "Seasonal maintenance assumption",
         "what_it_supports": "11_Availability_Outages for CHP and run-of-river hydro (no UMM coverage)",
         "url": "", "notes": "documented in docs/METHODOLOGY.md (§6) item 7"},
        {"source_id": "S31", "source_name": f"ENTSO-E Transparency Platform - UMM Hydro Pumped Storage {DATA_YEAR}",
         "what_it_supports": "10_FlexAvailability PSP_COO_PLATE_TAILLE availability",
         "url": "https://transparency.entsoe.eu/outage-domain/r2/unavailabilityOfGenerationUnits/show", "notes": f"{COUNTRY_CODE} pumped-storage outage events where available"},
        {"source_id": "S32", "source_name": "Battery storage technical availability assumption",
         "what_it_supports": "10_FlexAvailability BESS_1H/BESS_4H",
         "url": "", "notes": "Asset-level BESS availability is not published (restricted, item 5 in data-source map); flat 92% technical-availability assumption"},
        {"source_id": "S33", "source_name": "EV/heat-pump diurnal+seasonal flexibility assumption",
         "what_it_supports": "10_FlexAvailability EV_SMART_CHARGING / HEATPUMP_SHIFT",
         "url": "https://arxiv.org/abs/2409.18105", "notes": "Documented diurnal plug-in / heating-season pattern, no asset-level metering available (item 14 in data-source map)"},
        {"source_id": "S34", "source_name": "Industrial/commercial DR availability assumption",
         "what_it_supports": "10_FlexAvailability INDUSTRIAL_DR_FAST/SLOW, COMMERCIAL_HVAC",
         "url": "", "notes": "Weekday production/business-hours pattern with August industrial shutdown; aggregator-level data is restricted (item 14 in data-source map)"},
        {"source_id": "S35", "source_name": f"{COUNTRY_CODE} FNA data-source map",
         "what_it_supports": f"All sheets - canonical mapping of input categories to {COUNTRY_CODE}/EU sources",
         "url": "", "notes": "See docs/METHODOLOGY.md (§6)"},
    ]
    return pd.concat([sources, pd.DataFrame(new_rows)], ignore_index=True)


def refresh_res_portfolios(res: pd.DataFrame) -> pd.DataFrame:
    res = res.copy()
    res_mask = res["technology"].isin(["offshore_wind", "onshore_wind", "solar_pv"])
    res.loc[res_mask, "source_id"] = res.loc[res_mask, "source_id"].astype(str) + "/S26"
    res.loc[res_mask, "notes"] = (
        res.loc[res_mask, "notes"].astype(str)
        + f" Cross-checked against ENTSO-E {DATA_YEAR} actual generation for {COUNTRY_CODE} when building capacity-factor profiles."
    )
    return res


def refresh_flex_storage(flex: pd.DataFrame) -> pd.DataFrame:
    flex = flex.copy()
    mask = flex["flex_id"] == "PSP_COO_PLATE_TAILLE"
    flex.loc[mask, ["up_capacity_MW_2025", "down_capacity_MW_2025", "up_capacity_MW_2030", "down_capacity_MW_2030"]] = INSTALLED["hydro_psp"]
    flex.loc[mask, "source_id"] = "S31"
    flex.loc[mask, "data_quality"] = f"ENTSO-E {DATA_YEAR} installed Hydro Pumped Storage capacity ({INSTALLED['hydro_psp']:.1f} MW)"
    return flex


def refresh_control(control: pd.DataFrame) -> pd.DataFrame:
    control = control.copy()
    updates = {
        "time_resolution": (f"full_year_hourly_{N_HOURS}h_{DATA_YEAR}base", "category", f"{N_DAYS} calendar days x 24h, no clustering (D001..D{N_DAYS:03d})", "model design", "S25"),
        "entsoe_country_code": (COUNTRY_CODE, "code", "ENTSO-E country code used to build the full-year hourly profiles", "user input", "S25"),
        "entsoe_data_year": (DATA_YEAR, "year", "ENTSO-E data year used to build the full-year hourly profiles", "user input", "S25"),
        "scenario_id": (f"{COUNTRY_CODE}_FNA_v3.1_FULLYEAR{BUILD_YEAR}", "Text", f"Full-year hourly {COUNTRY_CODE} FNA input set ({N_HOURS} h, {DATA_YEAR} data year)", "assumption", "S0"),
    }
    for param, (value, unit, desc, dq, src) in updates.items():
        mask = control["parameter"] == param
        if mask.any():
            control.loc[mask, "value"] = value
            control.loc[mask, "unit"] = unit
            control.loc[mask, "description"] = desc
            control.loc[mask, "data_quality"] = dq
            control.loc[mask, "source_id"] = src
        else:
            new_row = {c: np.nan for c in control.columns}
            new_row.update({"parameter": param, "value": value, "unit": unit, "description": desc, "data_quality": dq, "source_id": src})
            control = pd.concat([control, pd.DataFrame([new_row])], ignore_index=True)
    return control


def refresh_readme(readme: pd.DataFrame) -> pd.DataFrame:
    extra = pd.DataFrame([
        {readme.columns[0]: "Full-year variant", readme.columns[1]:
         f"This workbook (v3.1 {COUNTRY_CODE} FullYear{BUILD_YEAR}) replaces clustered representative days with all {N_DAYS} days of "
         f"{DATA_YEAR} ({N_HOURS} hourly rows, D001..D{N_DAYS:03d}, weight_days=1 each). Demand, wind/solar capacity factors, "
         f"cross-border flows, day-ahead prices, forecast-error stats and generation outages are ENTSO-E {DATA_YEAR} "
         "actuals where available; flexible-resource availability and reserve dimensioning remain documented assumptions per "
         "docs/METHODOLOGY.md (§6). Known limitation: storage SOC cycling is still enforced per "
         "calendar day (same_day_pairs in io_excel.write_inc_files), not per year - multi-day storage dynamics "
         "are therefore not captured without a GAMS-side change."},
    ])
    return pd.concat([readme, extra], ignore_index=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_full_year_workbook(
    country: str,
    year: int,
    template: "Path | None" = None,
    data_year: "int | None" = None,
    do_fetch: bool = True,
) -> Path:
    """Generate ``data/inputs/excel/<CC>_FullYear<year>.xlsx`` for a country/year.

    Hourly series (demand/wind/solar/flows/prices) are taken from the ENTSO-E
    cache in ``data/inputs/raw_<cc><data_year>/`` - fetched first when
    ``do_fetch`` and the cache is missing. Structural sheets (fleet, flex, interconnectors,
    control) come from ``template`` (default: the current input workbook). The
    output's assumption sheets/rows are colour-coded for review.
    """

    cc = _configure(country, year, template, data_year)
    dy = int(data_year or year)

    if not SRC_WB.exists():
        raise FileNotFoundError(
            f"Template workbook not found: {SRC_WB}. Pass --template <workbook in data/inputs/excel/> "
            f"with the structural sheets (06_DispatchableBlocks, 07_RES_Portfolios, "
            f"09_FlexStorage, 04_Interconnectors, 01_Control)."
        )

    if do_fetch and not RAW.exists():
        from fna.inputs.fetch import fetch_country_year

        print(f"Fetching ENTSO-E {cc} {dy} into {RAW} ...")
        fetch_country_year(cc, dy, RAW)
    if not RAW.exists():
        raise FileNotFoundError(
            f"No cached ENTSO-E data at {RAW}. Run with fetching enabled (needs ENTSOE_API_KEY)."
        )

    _build()
    summary = highlight_assumptions(OUT_WB)
    print_attention_notice(summary)
    print(f"Done. Wrote {OUT_WB}")
    return OUT_WB


def _build() -> None:
    if SRC_WB.resolve() == OUT_WB.resolve():
        print(f"Using existing workbook in place: {OUT_WB}")
    else:
        print(f"Copying {SRC_WB} -> {OUT_WB}")
        shutil.copy(SRC_WB, OUT_WB)

    print("Loading cached ENTSO-E series...")
    load = load_aligned("load")
    res_actual = load_aligned("res_generation_actual")
    res_fc = load_aligned("res_generation_forecast")
    prices = load_aligned("day_ahead_prices")
    flows = load_aligned("cross_border_flows")
    nb_prices = load_aligned("neighbour_day_ahead_prices", required=False)

    cal = build_calendar()

    print("Reading static input sheets...")
    res_portfolios = pd.read_excel(SRC_WB, sheet_name="07_RES_Portfolios")
    flex = pd.read_excel(SRC_WB, sheet_name="09_FlexStorage")
    dispatchable = pd.read_excel(SRC_WB, sheet_name="06_DispatchableBlocks")
    borders = pd.read_excel(SRC_WB, sheet_name="04_Interconnectors")
    sources = pd.read_excel(SRC_WB, sheet_name="20_Sources")
    control = pd.read_excel(SRC_WB, sheet_name="01_Control")
    readme = pd.read_excel(SRC_WB, sheet_name="00_ReadMe")

    print(f"Building 02_RepHours / 03_RepDays ({N_HOURS} h, {N_DAYS} days)...")
    hours, days = build_hours_and_days(cal, load)

    print("Building 08_RES_CF_Profiles...")
    res_cf = build_res_cf(cal, res_actual, res_portfolios)

    print("Building 05_IntercoProfiles...")
    interco = build_interco_profiles(cal, flows, prices, nb_prices, borders)

    print("Building 12_Reserve_ForecastError...")
    reserve = build_reserve_forecast_error(cal, load, res_actual, res_fc)

    print("Building 11_Availability_Outages...")
    outages = build_availability_outages(cal, dispatchable)

    print("Building 10_FlexAvailability...")
    flex_avail = build_flex_availability(cal, flex)

    print("Refreshing static sheets...")
    res_portfolios_new = refresh_res_portfolios(res_portfolios)
    flex_new = refresh_flex_storage(flex)
    sources_new = refresh_sources(sources)
    control_new = refresh_control(control)
    readme_new = refresh_readme(readme)

    sheets = {
        "02_RepHours": hours,
        "03_RepDays": days,
        "08_RES_CF_Profiles": res_cf,
        "05_IntercoProfiles": interco,
        "12_Reserve_ForecastError": reserve,
        "11_Availability_Outages": outages,
        "10_FlexAvailability": flex_avail,
        "07_RES_Portfolios": res_portfolios_new,
        "09_FlexStorage": flex_new,
        "20_Sources": sources_new,
        "01_Control": control_new,
        "00_ReadMe": readme_new,
    }

    print("Writing sheets...")
    with pd.ExcelWriter(OUT_WB, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
            print(f"  {name}: {df.shape}")

    print("Applying header styling and sheet order...")
    wb = openpyxl.load_workbook(OUT_WB)
    for name in sheets:
        header_style(wb[name])
    wb._sheets.sort(key=lambda sheet: sheet_sort_key(sheet.title))
    wb.save(OUT_WB)


def main(argv: list[str] | None = None) -> None:
    """Command-line entry point for direct module execution."""

    parser = argparse.ArgumentParser(description="Build a full-year hourly FNA input workbook.")
    parser.add_argument("--country", required=True, help="ENTSO-E country code, e.g. BE, FR, NL.")
    parser.add_argument("--year", type=int, required=True, help="Workbook/scenario year to write in the output name.")
    parser.add_argument("--data-year", type=int, default=None, help="ENTSO-E data year to read/fetch. Defaults to --year.")
    parser.add_argument("--template", type=Path, default=None, help="Template workbook path. Defaults to the configured input workbook.")
    parser.add_argument("--fetch", action=argparse.BooleanOptionalAction, default=False, help="Fetch ENTSO-E data if the cache is missing.")
    args = parser.parse_args(argv)

    template = args.template
    if template and not template.is_absolute():
        template = EXCEL_DIR / template.name

    build_full_year_workbook(
        country=args.country,
        year=args.year,
        data_year=args.data_year,
        template=template,
        do_fetch=args.fetch,
    )


if __name__ == "__main__":
    main()
