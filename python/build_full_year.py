"""
build_full_year.py - full 8760-hour benchmark mode (v3).

The main model runs on 240 representative hours. To quantify the error that the
representative-day compression introduces, this script fetches one full year of
ENTSO-E hourly data and writes it into a *copy* of the workbook as chronological
hours, with weight_days = 1/24 per hour so weighted annual demand equals the true
annual demand.

Workflow:
    python build_full_year.py          # writes Belgium_FNA_ED_v2_full_year.xlsx
    EXCEL_FILENAME=Belgium_FNA_ED_v2_full_year.xlsx python main.py
    # then compare 30_Summary_Tables and 40-43 FNA sheets between the two runs.

Use on small test systems only: the GAMS MIP grows with the number of hours.
This reuses the ENTSO-E fetch and column contract from rep_days.py.
"""
from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config import EXCEL_FILENAME, PROJECT_ROOT
from io_excel import sort_openpyxl_sheets
from rep_days import (
    build_res_cf_profiles,
    fetch_hourly_profiles,
    load_config,
    read_workbook_tables,
    season_from_month,
    day_type_from_date,
    write_dataframe_to_sheet,
)


def main() -> None:
    cfg = load_config()
    out_path = cfg.excel_path.with_name(cfg.excel_path.stem.replace("_input_data", "") + "_full_year.xlsx")
    shutil.copyfile(cfg.excel_path, out_path)
    print(f"Building full-year benchmark workbook: {out_path.name}")

    from entsoe import EntsoePandasClient

    client = EntsoePandasClient(api_key=cfg.api_key)
    profiles = fetch_hourly_profiles(client, cfg)
    profiles = profiles[profiles.index.year == cfg.data_year].sort_index()

    tables = read_workbook_tables(cfg.excel_path)
    rep_hours_df = _full_year_hours(profiles, cfg)
    rep_days_df = _single_chronology_day(cfg)
    res_cf_df = _full_year_res_cf(profiles, tables["res"], cfg)

    _write(out_path, {
        "02_RepHours": rep_hours_df,
        "03_RepDays": rep_days_df,
        "08_RES_CF_Profiles": res_cf_df,
    })
    print(f"Done. {len(rep_hours_df)} hourly rows written. Run main.py with EXCEL_FILENAME={out_path.name}.")


def _full_year_hours(profiles: pd.DataFrame, cfg: Any) -> pd.DataFrame:
    """One chronological group of 8760 hours; weight 1/24 day per hour."""
    rows: list[dict[str, Any]] = []
    index = list(profiles.index)
    for i, ts in enumerate(index):
        nxt = index[i + 1] if i + 1 < len(index) else None
        rows.append({
            "time_id": ts.strftime("H%Y%m%d_%H"),
            "rep_day_id": "FULLYEAR",
            "hour": int(ts.hour),
            "next_time_id": index[i + 1].strftime("H%Y%m%d_%H") if nxt is not None else "",
            "season": season_from_month(ts.month),
            "day_type": day_type_from_date(ts),
            "weight_days": 1.0 / 24.0,
            "weight_hours": 1.0 / 24.0,
            "chronology_group": "FULLYEAR",
            f"gross_demand_MW_{cfg.target_year}": float(profiles.loc[ts, "load"]),
            "source_id": cfg.source_id,
            "data_quality": f"ENTSO-E {cfg.data_year} full year",
            "notes": "Full-year benchmark hour",
        })
    return pd.DataFrame(rows)


def _single_chronology_day(cfg: Any) -> pd.DataFrame:
    return pd.DataFrame([{
        "rep_day_id": "FULLYEAR",
        "description": f"Full year {cfg.data_year}",
        "season": "all",
        "day_type": "all",
        "weight_days": 365.0,
        "selection_reason": "Full 8760-hour benchmark",
        "probability_pct": 100.0,
        "source_id": cfg.source_id,
        "data_quality": "ENTSO-E full year",
    }])


def _full_year_res_cf(profiles: pd.DataFrame, res_portfolios: pd.DataFrame, cfg: Any) -> pd.DataFrame:
    """Hourly capacity factors for the whole year using workbook capacities."""
    capacity_col = next(
        (c for c in [f"capacity_mw_{cfg.target_year}", "capacity_mw"] if c in res_portfolios.columns),
        None,
    )
    if capacity_col is None:
        raise KeyError("No RES capacity column found for full-year build.")

    active = res_portfolios[pd.to_numeric(res_portfolios[capacity_col], errors="coerce").fillna(0.0) > 0.0]
    rows: list[dict[str, Any]] = []
    for ts in profiles.index:
        for _, res_row in active.iterrows():
            tech = str(res_row.get("technology", "")).lower()
            gen_col = "wind" if "wind" in tech else ("solar" if ("solar" in tech or "pv" in tech) else None)
            if gen_col is None:
                continue
            cap = float(res_row[capacity_col])
            cf = float(profiles.loc[ts, gen_col]) / cap if cap > 0 else 0.0
            rows.append({
                "time_id": ts.strftime("H%Y%m%d_%H"),
                "res_id": res_row["res_id"],
                "capacity_factor": min(max(cf, 0.0), 1.0),
                "availability_pct": cfg.availability_pct,
                "source_id": cfg.source_id,
                "data_quality": f"ENTSO-E {cfg.data_year} full year",
                "notes": "Full-year benchmark CF",
            })
    return pd.DataFrame(rows)


def _write(path: Path, tables: dict[str, pd.DataFrame]) -> None:
    wb = load_workbook(path)
    for sheet_name, df in tables.items():
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        write_dataframe_to_sheet(ws, df)
    sort_openpyxl_sheets(wb)
    wb.save(path)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
