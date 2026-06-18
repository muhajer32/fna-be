"""
main.py - Belgium FNA-ED/UC v2 unified workflow.

Runs either a deterministic single case or a Monte Carlo uncertainty run from
the same entry point. Excel/config controls the default mode; CLI flags can
override it for scripts and macros.
"""
from __future__ import annotations

import logging
import shutil
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook

_HERE = Path(__file__).resolve().parent
if str(_HERE) not in sys.path:
    sys.path.insert(0, str(_HERE))

from config import EXCEL_FILENAME, PATHS, PECD_COUNTRY_CODE, PECD_YEARS, PROJECT_ROOT, WIND_RESOURCE_IDS, output_excel_path
from io_excel import (
    apply_uncertainty_scenario,
    parse_csv_results,
    read_inputs,
    read_inputs_with_mc,
    read_uncertainty_params,
    write_mc_charts_to_excel,
    write_mc_results_to_excel,
    write_results,
)
from monte_carlo import MonteCarloScenarios, PECDReader, RepresentativeHourMap
from plot_results import generate_all, generate_mc_summary_charts
from run_gams import run_model

log = logging.getLogger("main")


def main() -> None:
    started_at = time.perf_counter()
    wb = _open_workbook()
    params = read_uncertainty_params(wb)
    run_mc = _use_monte_carlo(params)
    log_file = PROJECT_ROOT / "logs" / ("run_mc.log" if run_mc else "run.log")

    _clean_generated_outputs()
    _reset_output_workbook(wb)
    _configure_logging(log_file)

    try:
        if run_mc:
            run_monte_carlo(wb, params, started_at=started_at)
        else:
            run_optimisation(wb, started_at=started_at)
    finally:
        log.info("Total main.py runtime: %s", _format_elapsed(time.perf_counter() - started_at))


def run_optimisation(
    wb: Any | None = None,
    started_at: float | None = None,
    inputs: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
) -> dict[str, Any]:
    """Run the deterministic mean/base case.

    ``inputs``, ``paths`` and ``sheet_suffix`` let multi_year.py reuse this
    function for each target year: ``inputs`` overrides the workbook-read
    inputs (e.g. with a year-specific target_year and demand column), ``paths``
    overrides the global PATHS with a per-year folder, and ``sheet_suffix`` is
    appended to every output sheet name. Returns the GAMS results dict, with
    the extra ACER tables (from write_results) under "fna_tables".
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 deterministic run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    out_wb = _open_output_workbook(wb)
    paths = paths or PATHS
    status_cell = _safe_status_cell(wb)

    def status(msg: str) -> None:
        _set_status(status_cell, msg)
        log.info(msg)

    try:
        _ensure_directories(paths)
        if inputs is None:
            status("Reading v2 Excel inputs...")
            inputs = read_inputs(wb)

        status("Running GAMS 25.1 deterministic model...")
        results = run_model(inputs, {k: str(v) for k, v in paths.items()})

        status("Writing output tables...")
        extra_tables = write_results(
            out_wb,
            results,
            rep_hours=inputs["frames"].get("hours"),
            rep_days=inputs["frames"].get("days"),
            network_needs=inputs["frames"].get("network"),
            dso_zones=inputs["frames"].get("dso_zones"),
            flex_availability=inputs["frames"].get("flex_profiles"),
            flex=inputs["frames"].get("flex"),
            prequalification=inputs["frames"].get("prequalification"),
            control=inputs.get("control"),
            target_year=inputs.get("target_year"),
            input_frames=inputs["frames"],
            sheet_suffix=sheet_suffix,
            img_dir=Path(paths["img_dir"]),
        )
        results["fna_tables"] = extra_tables

        status("Generating FNA output charts...")
        generate_all(results=results, img_dir=Path(paths["img_dir"]), wb=out_wb)

        if save:
            out_wb.save(output_excel_path())
        status(f"Done: deterministic optimisation completed in {_elapsed_since(started_at)}.")
        return results
    except Exception as exc:
        status("Error: see logs/run.log and logs/gams_run.lst")
        log.exception("Run failed: %s", exc)
        raise


def run_postprocess(
    wb: Any | None = None,
    started_at: float | None = None,
    inputs: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
) -> dict[str, Any]:
    """Recompute the ACER FNA indicator / network-need sheets and charts from
    the CSV outputs of a *previous* GAMS run, without re-running GAMS.

    Used by the ``compute-fna-indicators`` and ``fine-tune`` CLI commands, so
    the (slow) optimisation does not have to be repeated just to re-derive the
    post-processed indicator tables.
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 post-processing run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    out_wb = _open_output_workbook(wb)
    paths = paths or PATHS

    out_dir = Path(paths["out_dir"])
    log.info("Reading existing GAMS CSV outputs from %s", out_dir)
    results = parse_csv_results(out_dir)

    if inputs is None:
        log.info("Reading v2 Excel inputs...")
        inputs = read_inputs(wb)

    log.info("Writing FNA indicator / network-need tables...")
    extra_tables = write_results(
        out_wb,
        results,
        rep_hours=inputs["frames"].get("hours"),
        rep_days=inputs["frames"].get("days"),
        network_needs=inputs["frames"].get("network"),
        dso_zones=inputs["frames"].get("dso_zones"),
        flex_availability=inputs["frames"].get("flex_profiles"),
        flex=inputs["frames"].get("flex"),
        prequalification=inputs["frames"].get("prequalification"),
        control=inputs.get("control"),
        target_year=inputs.get("target_year"),
        input_frames=inputs["frames"],
        sheet_suffix=sheet_suffix,
        img_dir=Path(paths["img_dir"]),
    )
    results["fna_tables"] = _collect_indicator_tables(results, inputs, extra_tables)

    log.info("Generating FNA output charts...")
    generate_all(results=results, img_dir=Path(paths["img_dir"]), wb=out_wb)

    if save:
        out_wb.save(output_excel_path())
    log.info("Done: post-processing completed in %s.", _elapsed_since(started_at))
    return results


def _collect_indicator_tables(
    results: dict[str, Any],
    inputs: dict[str, Any],
    extra_tables: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    """Merge the consolidated report tables (``extra_tables``, from
    ``write_results``) with the raw per-need indicator tables from
    ``build_fna_indicators`` and the DSO/TSO/Article-14 network tables, so
    CLI commands (``compute-fna-indicators``, ``fine-tune``, ``make-report``)
    have a single dict of named DataFrames to export/report on."""

    tables: dict[str, pd.DataFrame] = dict(extra_tables)
    frames = inputs["frames"]
    control = inputs.get("control") or {}

    try:
        from fna_indicators import build_fna_indicators

        tables.update(build_fna_indicators(
            results["residual"], frames.get("hours"), frames.get("days"),
            market_time_unit_minutes=float(control.get("market_time_unit_minutes", 60) or 60),
            flex_availability=frames.get("flex_profiles"), flex=frames.get("flex"),
            storage=results.get("storage"), prequalification=frames.get("prequalification"),
            target_year=inputs.get("target_year"),
        ))
    except Exception as exc:
        log.warning("Could not build per-need FNA indicator tables: %s", exc)

    try:
        from network_needs import compute_dso_needs, compute_fine_tuning_needs, compute_tso_needs

        tables["dso_needs"] = compute_dso_needs(results["residual"], frames.get("hours"), frames.get("dso_zones"))
        network_csv = results.get("network")
        if isinstance(network_csv, pd.DataFrame) and not network_csv.empty:
            tables["tso_needs"] = compute_tso_needs(network_csv, frames.get("network"), frames.get("hours"))
            tables["fine_tuning"] = compute_fine_tuning_needs(network_csv, frames.get("network"), frames.get("hours"))
    except Exception as exc:
        log.warning("Could not build DSO/TSO/Article-14 network-need tables: %s", exc)

    return tables


def run_monte_carlo(
    wb: Any | None = None,
    mc_params: dict[str, Any] | None = None,
    started_at: float | None = None,
    inputs_base: dict[str, Any] | None = None,
    paths: dict[str, Any] | None = None,
    sheet_suffix: str = "",
    save: bool = True,
) -> dict[int, dict]:
    """Run the full Monte Carlo workflow.

    ``inputs_base``, ``paths`` and ``sheet_suffix`` mirror run_optimisation:
    they let multi_year.py run the MC workflow once per target year, writing
    scenarios to a per-year folder and results to suffixed Excel sheets.
    Returns the per-scenario MC results dict.
    """

    log.info("=" * 70)
    log.info("Belgium FNA-ED/UC v2 Monte Carlo run started%s", f" (target year {sheet_suffix.lstrip('_')})" if sheet_suffix else "")
    wb = wb or _open_workbook()
    out_wb = _open_output_workbook(wb)
    mc_params = dict(mc_params or read_uncertainty_params(wb))
    paths = paths or PATHS
    status_cell = _safe_status_cell(wb)

    def status(message: str) -> None:
        _set_status(status_cell, message)
        log.info(message)

    try:
        _ensure_directories(paths)
        n_scenarios = int(mc_params["n_mc_scenarios"])
        max_workers = int(mc_params["max_parallel_workers"])
        seed = mc_params.get("seed_random")
        use_pecd = bool(mc_params.get("use_pecd_data"))

        if inputs_base is None:
            status("Reading base case inputs...")
            inputs_base = read_inputs_with_mc(wb, scenario_id=None)
        mc_params["pecd_target_year"] = int(mc_params.get("pecd_target_year") or inputs_base["target_year"])
        wind_portfolios = _wind_portfolios(inputs_base)
        _validate_wind_capacity(mc_params.get("wind_capacity_mw"), _wind_capacity_mw(inputs_base, wind_portfolios))

        status("Generating PECD load, wind and solar uncertainty scenarios...")
        mc_gen = MonteCarloScenarios(n_scenarios=n_scenarios, seed=seed)
        existing_scenarios = None
        rep_hours = _representative_hour_map(inputs_base)
        demand_pecd, solar_pecd, common_years = _load_pecd_load_solar(mc_params, inputs_base)
        if use_pecd:
            existing_scenarios = _generate_pecd_scenarios(mc_gen, mc_params, rep_hours, wind_portfolios, common_years)
        else:
            log.info("use_pecd_data is false; wind uses workbook mean/base profile.")

        uncertainty_scenarios = _add_pecd_load_solar_scenarios(
            mc_gen,
            inputs_base,
            rep_hours,
            demand_pecd,
            solar_pecd,
            existing_scenarios=existing_scenarios,
        )

        status(f"Running {n_scenarios} scenarios...")
        mc_results = _run_scenarios(
            n_scenarios,
            max_workers,
            inputs_base,
            uncertainty_scenarios,
            paths=paths,
        )
        if not mc_results:
            raise RuntimeError("All scenarios failed.")

        status("Writing Monte Carlo results...")
        write_mc_results_to_excel(out_wb, mc_results, requested_scenarios=n_scenarios, sheet_suffix=sheet_suffix)

        status("Generating charts...")
        chart_paths = _generate_charts(mc_results, uncertainty_scenarios, n_scenarios, max_workers, img_dir=Path(paths["img_dir"]))
        write_mc_charts_to_excel(out_wb, chart_paths, len(mc_results), n_scenarios, sheet_suffix=sheet_suffix)

        if save:
            out_wb.save(output_excel_path())
        status(f"Done: Monte Carlo completed with {len(mc_results)} successful scenarios in {_elapsed_since(started_at)}")
        log.info("Base case cost: %.0f EUR", _get_base_cost(mc_results))
        log.info("Cost std dev: %.0f EUR", _get_cost_std(mc_results))
        log.info("Results written to %s", output_excel_path())
        return mc_results
    except Exception as exc:
        status("Error: see logs/run_mc.log")
        log.exception("Monte Carlo failed: %s", exc)
        raise


def run_single_scenario(
    scenario_id: int,
    inputs_base: dict,
    uncertainty_scenarios: dict[int, pd.DataFrame],
    paths: dict[str, str],
    inc_dir_scenario: Path,
    out_dir_scenario: Path,
) -> dict:
    """Run one scenario in a worker process without opening Excel."""

    try:
        log.info("[Scenario %s] Starting", scenario_id)
        inputs = apply_uncertainty_scenario(inputs_base, scenario_id, uncertainty_scenarios)

        scenario_paths = dict(paths)
        scenario_paths["inc_dir"] = str(inc_dir_scenario)
        scenario_paths["out_dir"] = str(out_dir_scenario)
        scenario_paths["log_file"] = str(out_dir_scenario / "gams_run.log")
        results = run_model(inputs, scenario_paths)

        cost = _result_indicator(results, "total_cost", np.nan)
        log.info("[Scenario %s] Complete (cost: %.0f EUR)", scenario_id, cost)
        return {"scenario_id": scenario_id, "results": results, "cost": cost, "error": None}
    except Exception as exc:
        log.error("[Scenario %s] Failed: %s", scenario_id, exc, exc_info=True)
        return {"scenario_id": scenario_id, "results": None, "cost": np.nan, "error": str(exc)}


def _use_monte_carlo(params: dict[str, Any]) -> bool:
    return bool(params.get("run_monte_carlo", False))


def _configure_logging(log_file: Path) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        handlers=[logging.FileHandler(log_file, mode="w", encoding="utf-8"), logging.StreamHandler(sys.stdout)],
        force=True,
    )


def _open_workbook():
    """Load the input workbook headlessly with openpyxl (read-only).

    No live Excel/COM is required, so the model runs on any platform including
    headless Linux containers. ``data_only=True`` reads cached cell values; the
    FNA workbooks contain plain values (no live formulas) so this is exact.
    """

    wb_path = PROJECT_ROOT / "excel" / EXCEL_FILENAME
    return load_workbook(wb_path, read_only=True, data_only=True)


def _open_output_workbook(input_wb: Any = None):
    """Open (or create) the separate ``<input stem>-output.xlsx`` workbook used
    for all results/charts. Loads the existing file when present (so post-process
    runs update it in place); otherwise starts a clean, empty workbook."""

    out_path = output_excel_path()
    if out_path.exists():
        return load_workbook(out_path)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty "Sheet"; real sheets are added on write
    return wb


def _ensure_directories(paths: dict[str, Any] | None = None) -> None:
    paths = paths or PATHS
    for path in [paths["out_dir"], paths["img_dir"], paths["log_file"].parent, paths["inc_dir"]]:
        Path(path).mkdir(parents=True, exist_ok=True)


def _reset_output_workbook(input_wb: Any = None) -> None:
    """Delete the previous ``<input stem>-output.xlsx`` so each successful run
    starts from a clean workbook."""

    out_path = output_excel_path()
    if out_path.exists():
        try:
            out_path.unlink()
        except OSError as exc:
            log.warning("Could not remove previous output workbook %s: %s", out_path, exc)


def _clean_generated_outputs(paths: dict[str, Any] | None = None) -> None:
    """Delete generated scenarios, images and old logs before each run, so a
    new run never reuses a previous run's results."""

    paths = paths or PATHS
    for root in [Path(paths["inc_dir"]), Path(paths["out_dir"])]:
        if not root.exists():
            continue
        for path in root.glob("scenario_*"):
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)

    img_dir = Path(paths["img_dir"])
    if img_dir.exists():
        for image in img_dir.glob("*.png"):
            try:
                image.unlink()
            except OSError:
                pass

    logs_dir = Path(paths["log_file"]).parent
    if logs_dir.exists():
        for pattern in ["*.log", "*.lst"]:
            for log_path in logs_dir.glob(pattern):
                try:
                    log_path.unlink()
                except OSError:
                    pass


def _generate_pecd_scenarios(
    mc_gen: MonteCarloScenarios,
    mc_params: dict,
    rep_hours: list[RepresentativeHourMap],
    wind_portfolios: pd.DataFrame,
    common_years: list[int],
) -> dict[int, pd.DataFrame]:
    pecd_dir = _project_path(mc_params.get("pecd_data_dir") or PROJECT_ROOT / "data" / "pecd")
    reader = PECDReader(pecd_dir)
    target_year = int(mc_params["pecd_target_year"])

    frames = []
    for wind_type in sorted(wind_portfolios["wind_type"].unique()):
        frames.append(reader.read_wind_time_series(PECD_COUNTRY_CODE, wind_type, target_year, weather_years=common_years))

    pecd_cf = pd.concat(frames, ignore_index=True)
    return mc_gen.generate_from_pecd(pecd_cf, rep_hours, wind_portfolios)


def _load_pecd_load_solar(
    mc_params: dict,
    inputs_base: dict,
) -> tuple[pd.DataFrame, pd.DataFrame, list[int]]:
    pecd_dir = _project_path(mc_params.get("pecd_data_dir") or PROJECT_ROOT / "data" / "pecd")
    reader = PECDReader(pecd_dir)
    requested_years = PECD_YEARS or None
    target_year = int(mc_params.get("pecd_target_year") or inputs_base["target_year"])
    demand_pecd = reader.read_demand_time_series(PECD_COUNTRY_CODE, target_year, weather_years=requested_years)
    solar_pecd = reader.read_solar_time_series(PECD_COUNTRY_CODE, target_year, weather_years=requested_years)
    demand_years = set(int(y) for y in demand_pecd["weather_year"].dropna().unique())
    solar_years = set(int(y) for y in solar_pecd["weather_year"].dropna().unique())
    common_years = sorted(demand_years & solar_years)
    if not common_years:
        raise ValueError("No common PECD weather years between demand and solar data.")
    log.info("Using %d common PECD demand/solar weather years: %s-%s", len(common_years), common_years[0], common_years[-1])
    return demand_pecd, solar_pecd, common_years


def _add_pecd_load_solar_scenarios(
    mc_gen: MonteCarloScenarios,
    inputs_base: dict,
    rep_hours: list[RepresentativeHourMap],
    demand_pecd: pd.DataFrame,
    solar_pecd: pd.DataFrame,
    existing_scenarios: dict[int, pd.DataFrame] | None,
) -> dict[int, pd.DataFrame]:
    return mc_gen.add_pecd_load_solar(
        inputs_base,
        rep_hours,
        demand_pecd,
        solar_pecd,
        existing_scenarios=existing_scenarios,
    )


def _project_path(path_value: object) -> Path:
    path = Path(str(path_value)).expanduser()
    if path.is_absolute():
        return path
    return PROJECT_ROOT / path


def _run_scenarios(
    n_scenarios: int,
    max_workers: int,
    inputs_base: dict,
    uncertainty_scenarios: dict[int, pd.DataFrame],
    paths: dict[str, Any] | None = None,
) -> dict[int, dict]:
    paths = paths or PATHS
    scenario_dirs = {}
    for scenario_id in range(n_scenarios):
        inc_dir = Path(paths["inc_dir"]) / f"scenario_{scenario_id}"
        out_dir = Path(paths["out_dir"]) / f"scenario_{scenario_id}"
        scenario_dirs[scenario_id] = (inc_dir, out_dir)

    paths_dict = {key: str(value) for key, value in paths.items()}
    mc_results: dict[int, dict] = {}

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for scenario_id in range(n_scenarios):
            inc_dir, out_dir = scenario_dirs[scenario_id]
            future = executor.submit(
                run_single_scenario,
                scenario_id,
                inputs_base,
                uncertainty_scenarios,
                paths_dict,
                inc_dir,
                out_dir,
            )
            futures[future] = scenario_id

        completed = 0
        for future in as_completed(futures):
            result = future.result()
            scenario_id = result["scenario_id"]
            if result["error"] is None:
                mc_results[scenario_id] = result["results"]
                completed += 1
                log.info("Scenarios completed: %d/%d", completed, n_scenarios)
            else:
                log.error("Scenario %s failed: %s", scenario_id, result["error"])

    log.info("Completed %d/%d scenarios successfully", len(mc_results), n_scenarios)
    return mc_results


def _generate_charts(
    mc_results: dict[int, dict],
    uncertainty_scenarios: dict[int, pd.DataFrame],
    n_scenarios: int,
    max_workers: int,
    img_dir: Path | None = None,
) -> dict[str, Path]:
    return generate_mc_summary_charts(mc_results, uncertainty_scenarios, n_scenarios, max_workers, img_dir or PATHS["img_dir"])


def _representative_hour_map(inputs: dict) -> list[RepresentativeHourMap]:
    hours = inputs["frames"]["hours"].copy()
    days = inputs["frames"]["days"].copy()
    required_hours = {"time_id", "rep_day_id", "hour"}
    required_days = {"rep_day_id", "description"}
    if not required_hours.issubset(hours.columns):
        raise KeyError(f"02_RepHours missing columns: {sorted(required_hours - set(hours.columns))}")
    if not required_days.issubset(days.columns):
        raise KeyError(f"03_RepDays missing columns: {sorted(required_days - set(days.columns))}")

    rep_dates = {}
    for row in days.itertuples(index=False):
        date = _parse_representative_date(getattr(row, "description"))
        if date is not None:
            rep_dates[str(getattr(row, "rep_day_id"))] = date

    if not rep_dates:
        raise ValueError("Could not parse representative dates from 03_RepDays.description. Run rep_days.py first.")

    mapping = []
    for row in hours.itertuples(index=False):
        rep_day_id = str(getattr(row, "rep_day_id"))
        if rep_day_id not in rep_dates:
            raise ValueError(f"No representative date found for {rep_day_id}")
        date = rep_dates[rep_day_id]
        mapping.append(
            RepresentativeHourMap(
                time_id=str(getattr(row, "time_id")),
                month=int(date.month),
                day=int(date.day),
                hour=int(float(getattr(row, "hour"))),
            )
        )
    return mapping


def _parse_representative_date(text: object) -> pd.Timestamp | None:
    match = pd.Series([str(text)]).str.extract(r"((?:19|20)\d{2}-\d{2}-\d{2})").iloc[0, 0]
    if pd.isna(match):
        return None
    return pd.Timestamp(match)


def _wind_portfolios(inputs: dict) -> pd.DataFrame:
    res = inputs["frames"]["res"].copy()
    required = {"res_id", "technology"}
    if not required.issubset(res.columns):
        raise KeyError(f"07_RES_Portfolios missing columns: {sorted(required - set(res.columns))}")

    if WIND_RESOURCE_IDS:
        wind = res[res["res_id"].astype(str).isin(WIND_RESOURCE_IDS)].copy()
    else:
        wind = res[res["technology"].astype(str).str.lower().str.contains("wind", na=False)].copy()

    if wind.empty:
        raise ValueError("No wind portfolios found in 07_RES_Portfolios.")

    wind["wind_type"] = wind["technology"].map(_wind_type_from_technology)
    if wind["wind_type"].isna().any():
        bad = wind.loc[wind["wind_type"].isna(), ["res_id", "technology"]].to_dict("records")
        raise ValueError(f"Cannot classify wind portfolios as onshore/offshore: {bad}")
    return wind[["res_id", "technology", "wind_type"]].copy()


def _wind_type_from_technology(value: object) -> str | None:
    text = str(value).lower()
    if "offshore" in text:
        return "offshore"
    if "onshore" in text or "wind" in text:
        return "onshore"
    return None


def _wind_capacity_mw(inputs: dict, wind_portfolios: pd.DataFrame) -> float:
    target_year = int(inputs["target_year"])
    capacity_col = f"capacity_mw_{target_year}"
    res = inputs["frames"]["res"].copy()
    if capacity_col not in res.columns:
        raise KeyError(f"Missing {capacity_col} in 07_RES_Portfolios.")

    wind_ids = set(wind_portfolios["res_id"].astype(str))
    capacity = pd.to_numeric(
        res.loc[res["res_id"].astype(str).isin(wind_ids), capacity_col],
        errors="coerce",
    ).fillna(0.0).sum()
    if capacity <= 0:
        raise ValueError(f"Derived wind capacity is {capacity}; check 07_RES_Portfolios.")
    return float(capacity)


def _validate_wind_capacity(configured_capacity: object, derived_capacity: float) -> None:
    if configured_capacity is None or pd.isna(configured_capacity):
        log.info("Derived wind capacity from 07_RES_Portfolios: %.2f MW", derived_capacity)
        return

    configured = float(configured_capacity)
    if configured <= 0:
        raise ValueError("wind_capacity_mw must be positive if supplied.")

    rel_diff = abs(configured - derived_capacity) / derived_capacity
    if rel_diff > 0.01:
        raise ValueError(
            f"wind_capacity_mw={configured:.2f} does not match 07_RES_Portfolios "
            f"wind capacity={derived_capacity:.2f} MW."
        )
    log.info("Validated configured wind capacity %.2f MW", configured)


def _safe_status_cell(wb):
    # Headless runs have no live Excel to surface progress in; status is logged
    # instead (see the status() closures), so there is no status cell to write.
    return None


def _set_status(cell: Any, message: str) -> None:
    # No-op: progress is reported through logging in headless mode.
    return None


def _elapsed_since(started_at: float | None) -> str:
    if started_at is None:
        return "unknown time"
    return _format_elapsed(time.perf_counter() - started_at)


def _format_elapsed(seconds: float) -> str:
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def _get_base_cost(mc_results: dict[int, dict]) -> float:
    if 0 in mc_results:
        return _result_indicator(mc_results[0], "total_cost", np.nan)
    return np.nan


def _get_cost_std(mc_results: dict[int, dict]) -> float:
    costs = []
    for result in mc_results.values():
        cost = _result_indicator(result, "total_cost")
        if cost is not None and not pd.isna(cost):
            costs.append(float(cost))
    return float(np.std(costs)) if len(costs) > 1 else 0.0


def _result_indicator(result: dict, metric: str, default: object = None) -> object:
    indicators = result.get("indicators", {})
    if isinstance(indicators, dict):
        return indicators.get(metric, default)
    if isinstance(indicators, pd.DataFrame) and {"metric", "value"}.issubset(indicators.columns):
        rows = indicators[indicators["metric"].astype(str).str.strip().eq(metric)]
        if not rows.empty:
            return pd.to_numeric(rows.iloc[0]["value"], errors="coerce")
    return default


if __name__ == "__main__":
    main()
