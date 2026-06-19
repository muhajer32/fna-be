"""
build_full_year_workbook.py - build a full-year (8760 h, 2023) Belgium FNA
input workbook from cached ENTSO-E data (data/raw_be2023/).

This produces excel/Belgium_FNA_v3.1_FullYear2023_input_data.xlsx, a sibling
of the representative-day workbook (Belgium_FNA_ED_v2_input_data.xlsx) with
the SAME sheet/column schema, but:

  - 02_RepHours / 03_RepDays: 365 individual calendar days (D001..D365),
    weight_days=1 each, instead of 10 clustered representative days. This is
    "no clustering" while staying inside the existing rep_day_id/weight_days
    contract (see io_excel.write_inc_files), so the GAMS model can run on it
    unmodified -- with the caveat noted in 00_ReadMe that storage SOC cycles
    are currently enforced per calendar day (same_day_pairs), not per year.
  - 08_RES_CF_Profiles, 05_IntercoProfiles, 10_FlexAvailability,
    11_Availability_Outages, 12_Reserve_ForecastError: re-keyed to the same
    8760 time_ids, populated with real ENTSO-E 2023 series where available
    (demand, wind/solar CF, cross-border flows, day-ahead prices, forecast
    errors, generation outages) and documented robust assumptions elsewhere
    (flexible-resource availability, reserve dimensioning) per
    docs/METHODOLOGY.md (§6).
  - 04/06/07/09/13/13b/17/20/01_Control/00_ReadMe: lightly refreshed
    source_id/data_quality/notes; numeric assumptions only changed where a
    real 2023 ENTSO-E figure is a clear improvement (documented per row).

Run after fetch_be_2023_data.py has populated data/raw_be2023/.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill


from fna_be.config import EXCEL_FILENAME, PROJECT_ROOT  # noqa: E402

RAW = PROJECT_ROOT / "data" / "raw_be2023"
SRC_WB = PROJECT_ROOT / "excel" / EXCEL_FILENAME
OUT_WB = PROJECT_ROOT / "excel" / "Belgium_FNA_v3.1_FullYear2023_input_data.xlsx"

N_DAYS = 365
N_HOURS = N_DAYS * 24  # 8760
INDEX = pd.date_range("2023-01-01", periods=N_HOURS, freq="h", tz="UTC")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

INSTALLED = {
    "nuclear": 4937.0,
    "fossil_gas": 6987.62,
    "biomass": 730.34,
    "waste": 383.64,
    "wind_offshore": 2262.1,
    "wind_onshore": 3053.18,
    "solar": 6474.96,
    "hydro_psp": 1308.0,
    "hydro_ror": 185.51,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_aligned(name: str) -> pd.DataFrame:
    df = pd.read_csv(RAW / f"{name}.csv", index_col=0)
    df.index = pd.to_datetime(df.index, utc=True)
    df = df[~df.index.duplicated(keep="first")]
    df = df.reindex(INDEX)
    df = df.ffill().bfill()
    return df


def season_from_month(month: int) -> str:
    if month in (12, 1, 2):
        return "winter"
    if month in (3, 4, 5):
        return "spring"
    if month in (6, 7, 8):
        return "summer"
    return "autumn"


def day_type_from_date(ts: pd.Timestamp) -> str:
    return "weekday" if ts.weekday() < 5 else "weekend"


def build_calendar() -> pd.DataFrame:
    cal = pd.DataFrame(index=INDEX)
    day_no = (np.arange(N_HOURS) // 24) + 1
    hour = np.arange(N_HOURS) % 24
    cal["day_no"] = day_no
    cal["hour"] = hour
    cal["rep_day_id"] = [f"D{d:03d}" for d in day_no]
    cal["time_id"] = [f"D{d:03d}_H{h:02d}" for d, h in zip(day_no, hour)]
    cal["next_time_id"] = [
        f"D{d:03d}_H{(h + 1) % 24:02d}" for d, h in zip(day_no, hour)
    ]
    cal["season"] = [season_from_month(ts.month) for ts in INDEX]
    cal["day_type"] = [day_type_from_date(ts) for ts in INDEX]
    cal["date"] = INDEX.date
    return cal


def umm_availability(plant_type: str, capacity_mw: float) -> pd.Series:
    """Hourly availability_pct from real ENTSO-E 2023 UMM outage events."""

    df = pd.read_csv(RAW / "generation_unavailability.csv")
    sub = df[(df["plant_type"] == plant_type) & (df["docstatus"].isna())].copy()
    unavail = pd.Series(0.0, index=INDEX)
    if not sub.empty:
        sub["start"] = pd.to_datetime(sub["start"], utc=True)
        sub["end"] = pd.to_datetime(sub["end"], utc=True)
        sub["unavail_mw"] = (sub["nominal_power"] - sub["avail_qty"]).clip(lower=0.0)
        for _, row in sub.iterrows():
            mask = (INDEX >= row["start"]) & (INDEX < row["end"])
            unavail.loc[mask] += row["unavail_mw"]
    avail_pct = 1.0 - (unavail / capacity_mw).clip(lower=0.0, upper=0.95)
    return avail_pct


def header_style(ws) -> None:
    n_cols = ws.max_column
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def sheet_sort_key(name: str) -> tuple[int, str]:
    import re

    match = re.match(r"^(\d+)", str(name))
    if match:
        return int(match.group(1)), str(name).lower()
    return 999, str(name).lower()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_hours_and_days(cal: pd.DataFrame, load: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    hours = pd.DataFrame({
        "time_id": cal["time_id"],
        "rep_day_id": cal["rep_day_id"],
        "hour": cal["hour"],
        "next_time_id": cal["next_time_id"],
        "season": cal["season"],
        "day_type": cal["day_type"],
        "weight_days": 1,
        "weight_hours": 1,
        "chronology_group": cal["rep_day_id"],
        "gross_demand_MW_2025": load["load_actual"].to_numpy(),
        "source_id": "S25",
        "data_quality": "historical (ENTSO-E 2023 actual)",
        "notes": "Full calendar year 2023, no clustering; used as 2025 demand-level proxy.",
    })

    days_grp = cal.groupby("rep_day_id", sort=False).agg(
        date=("date", "first"), season=("season", "first"), day_type=("day_type", "first")
    ).reset_index()
    days = pd.DataFrame({
        "rep_day_id": days_grp["rep_day_id"],
        "description": [f"Calendar date {d}" for d in days_grp["date"]],
        "season": days_grp["season"],
        "day_type": days_grp["day_type"],
        "weight_days": 1,
        "selection_reason": "Actual calendar day - no clustering (full-year hourly dataset)",
        "probability_pct": 100.0 / N_DAYS,
        "source_id": "S25",
        "data_quality": "historical (ENTSO-E 2023 actual)",
    })
    return hours, days


def build_res_cf(cal: pd.DataFrame, res_actual: pd.DataFrame, res_portfolios: pd.DataFrame) -> pd.DataFrame:
    cf_wind_offshore = (res_actual["wind_offshore"] / INSTALLED["wind_offshore"]).clip(0, 1)
    cf_wind_onshore = (res_actual["wind_onshore"] / INSTALLED["wind_onshore"]).clip(0, 1)
    cf_solar = (res_actual["solar"] / INSTALLED["solar"]).clip(0, 1)

    cf_map = {
        "offshore_wind": (cf_wind_offshore, "ENTSO-E 2023 actual generation / ENTSO-E installed offshore wind capacity (2262.1 MW)"),
        "onshore_wind": (cf_wind_onshore, "ENTSO-E 2023 actual generation / ENTSO-E installed onshore wind capacity (3053.18 MW); single national profile applied to all onshore regions"),
        "solar_pv": (cf_solar, "ENTSO-E 2023 actual generation / ENTSO-E installed solar capacity (6474.96 MW); single national profile applied to all PV segments"),
    }

    avail_offshore = umm_availability("Wind Offshore", INSTALLED["wind_offshore"])

    rows = []
    for _, res in res_portfolios.iterrows():
        tech = str(res["technology"]).strip().lower()
        if tech not in cf_map:
            continue
        cf_series, note = cf_map[tech]
        avail = avail_offshore if tech == "offshore_wind" else pd.Series(0.98, index=INDEX)
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "res_id": res["res_id"],
            "capacity_factor": cf_series.to_numpy(),
            "availability_pct": avail.to_numpy(),
            "source_id": "S26",
            "data_quality": "historical (ENTSO-E 2023 actual)",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


def build_interco_profiles(cal: pd.DataFrame, flows: pd.DataFrame, prices: pd.DataFrame, nb_prices: pd.DataFrame, borders: pd.DataFrame) -> pd.DataFrame:
    be_price = prices["price_eur_mwh"]
    price_for_border = {
        "BE_FR": nb_prices.get("FR", be_price),
        "BE_NL": nb_prices.get("NL", be_price),
        "BE_DE": nb_prices.get("DE", be_price),
        "BE_UK": be_price,  # GB day-ahead price unavailable via this API; fall back to BE price
        "BE_LU": nb_prices.get("DE", be_price),  # Luxembourg is part of the DE/LU bidding zone
    }
    cap_by_border = borders.set_index("border_id")[["import_capacity_MW_2025", "export_capacity_MW_2025"]]

    rows = []
    for border_id in borders["border_id"]:
        exp = flows[f"{border_id}_export"]
        imp = flows[f"{border_id}_import"]
        net = imp - exp
        import_cap = float(cap_by_border.loc[border_id, "import_capacity_MW_2025"])
        export_cap = float(cap_by_border.loc[border_id, "export_capacity_MW_2025"])
        # Realised 2023 ENTSO-E physical flows occasionally exceed the
        # workbook's assumed 2025 NTC (e.g. BE_FR loop-flow effects); clip to
        # the assumed line capacity so fixed flows stay model-feasible.
        fixed_import = net.clip(lower=0.0, upper=import_cap)
        fixed_export = (-net).clip(lower=0.0, upper=export_cap)
        nb_price = price_for_border.get(border_id, be_price)
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "border_id": border_id,
            "availability_pct": 0.95,
            "fixed_import_MW": fixed_import.to_numpy(),
            "fixed_export_MW": fixed_export.to_numpy(),
            "import_price_EUR_MWh": nb_price.to_numpy(),
            "export_price_EUR_MWh": be_price.to_numpy(),
            "source_id": "S27",
            "data_quality": "historical (ENTSO-E 2023 actual flows/prices)" if border_id != "BE_UK" else "historical flows (ENTSO-E 2023); price = BE day-ahead (GB price unavailable)",
            "notes": "fixed_* = realised 2023 net flow, clipped to the 2025 import/export capacity assumption (BE_FR/BE_NL realised flows occasionally exceeded 3000 MW, likely loop-flow/PST effects); availability_pct = tradable headroom assumption (model can still trade up to capacity x availability_pct beyond the fixed flow).",
        }))
    return pd.concat(rows, ignore_index=True)


def build_reserve_forecast_error(cal: pd.DataFrame, load: pd.DataFrame, res_actual: pd.DataFrame, res_fc: pd.DataFrame) -> pd.DataFrame:
    season = cal["season"].to_numpy()
    hour = cal["hour"].to_numpy()

    load_err = (load["load_actual"] - load["load_forecast"]) / load["load_actual"]
    wind_actual = res_actual["wind_onshore"] + res_actual["wind_offshore"]
    wind_fc = res_fc["wind_onshore_fc"] + res_fc["wind_offshore_fc"]
    wind_cap = INSTALLED["wind_onshore"] + INSTALLED["wind_offshore"]
    wind_err = (wind_actual - wind_fc) / wind_cap
    solar_err = (res_actual["solar"] - res_fc["solar_fc"]) / INSTALLED["solar"]

    df = pd.DataFrame({"season": season, "load_err": load_err.to_numpy(), "wind_err": wind_err.to_numpy(), "solar_err": solar_err.to_numpy()})
    season_stats = df.groupby("season").agg(
        load_pct=("load_err", lambda s: float(s.std())),
        wind_pct=("wind_err", lambda s: float(s.std())),
        solar_pct=("solar_err", lambda s: float(s.std())),
    )

    load_pct = np.array([season_stats.loc[s, "load_pct"] for s in season])
    wind_pct = np.array([season_stats.loc[s, "wind_pct"] for s in season])
    # Solar forecast error only matters during daylight hours; keep it ~0 overnight.
    daylight = (hour >= 7) & (hour <= 18)
    solar_pct = np.where(daylight, [season_stats.loc[s, "solar_pct"] for s in season], 0.005)

    # Reserve dimensioning (FCR/aFRR/mFRR): kept as the existing Elia-informed
    # seasonal/diurnal pattern (S20/S21), extended from 10 representative days
    # to all 365 days. Winter evening peak (17-21h) gets the elevated band.
    winter_evening = (season == "winter") & (hour >= 17) & (hour <= 21)
    winter = season == "winter"

    fcr = np.full(N_HOURS, 90.0)
    afrr_up = np.where(winter_evening, 330.0, np.where(winter, 300.0, 250.0))
    afrr_dn = np.where(winter, 310.0, 250.0)
    mfrr_up = np.where(winter_evening, 1200.0, np.where(winter, 1000.0, 900.0))
    mfrr_dn = np.where(winter, 950.0, 750.0)
    largest_outage = np.where(winter_evening, 1000.0, 250.0)

    return pd.DataFrame({
        "time_id": cal["time_id"],
        "FCR_MW": fcr,
        "aFRR_up_MW": afrr_up,
        "aFRR_down_MW": afrr_dn,
        "mFRR_up_MW": mfrr_up,
        "mFRR_down_MW": mfrr_dn,
        "load_forecast_error_pct": load_pct,
        "wind_forecast_error_pct": wind_pct,
        "solar_forecast_error_pct": solar_pct,
        "largest_unit_outage_stress_MW": largest_outage,
        "source_id": "S20/S21/S28",
        "data_quality": "load/wind/solar error = empirical std (ENTSO-E 2023 forecast vs actual, per season); reserve volumes = Elia-informed seasonal/diurnal assumption",
        "notes": "Forecast-error pct = std((actual-forecast)/reference) per season, reference=actual load resp. installed wind/solar capacity. Reserve MW pattern unchanged from v2 (S20/S21), now applied to every day of the year.",
    })


def build_availability_outages(cal: pd.DataFrame, dispatchable: pd.DataFrame) -> pd.DataFrame:
    nuclear_avail = umm_availability("Nuclear", INSTALLED["nuclear"])
    gas_avail = umm_availability("Fossil Gas", INSTALLED["fossil_gas"])
    biomass_avail = umm_availability("Biomass", INSTALLED["biomass"])

    season = cal["season"]
    # No UMM coverage for run-of-river hydro or industrial CHP: keep a mild,
    # documented seasonal assumption (lower hydro availability in low-flow
    # summer; CHP maintenance concentrated in shoulder seasons).
    hydro_avail = season.map({"winter": 0.97, "spring": 0.95, "summer": 0.85, "autumn": 0.95}).astype(float)
    chp_avail = season.map({"winter": 0.95, "spring": 0.88, "summer": 0.92, "autumn": 0.88}).astype(float)

    profiles = {
        "NUC_EXIST": (nuclear_avail, "S29", "ENTSO-E 2023 UMM (REMIT) outage events for Belgian nuclear units"),
        "CCGT_EXIST": (gas_avail, "S29", "ENTSO-E 2023 UMM outage events for Belgian Fossil Gas units"),
        "CCGT_NEW": (gas_avail, "S29", "Same UMM-derived profile as CCGT_EXIST (no separate unit-level split available)"),
        "OCGT_PEAKER": (gas_avail, "S29", "Same UMM-derived profile as CCGT_EXIST (no separate unit-level split available)"),
        "CHP_INDUSTRIAL": (chp_avail, "S30", "No UMM coverage for industrial CHP; seasonal maintenance assumption"),
        "BIOMASS_WASTE": (biomass_avail, "S29", "ENTSO-E 2023 UMM outage events for Belgian Biomass units"),
        "HYDRO_RUN": (hydro_avail, "S30", "No UMM coverage for small hydro; seasonal low-flow assumption"),
    }

    rows = []
    for _, unit in dispatchable.iterrows():
        ugid = unit["unit_group_id"]
        if ugid not in profiles:
            continue
        avail, source_id, note = profiles[ugid]
        forced = (1.0 - avail).clip(lower=0.0)
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "unit_group_id": ugid,
            "planned_outage_pct": 0.0,
            "forced_outage_rate_pct": forced.to_numpy(),
            "availability_pct": avail.to_numpy(),
            "largest_unit_outage_MW": float(unit["capacity_per_block_MW_2025"]),
            "source_id": source_id,
            "data_quality": "historical (ENTSO-E 2023 UMM)" if source_id == "S29" else "seasonal assumption",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


def build_flex_availability(cal: pd.DataFrame, flex: pd.DataFrame) -> pd.DataFrame:
    hour = cal["hour"].to_numpy()
    weekday = (cal["day_type"] == "weekday").to_numpy()
    season = cal["season"].to_numpy()
    month = np.array([ts.month for ts in INDEX])

    psp_avail = umm_availability("Hydro Pumped Storage", INSTALLED["hydro_psp"])

    def diurnal(peak_hours, base, peak):
        return np.where(np.isin(hour, peak_hours), peak, base)

    # EV smart charging: cars plugged in mainly overnight + workplace midday;
    # discharge/V2G (up) availability lower than charge-shift (down).
    ev_down = diurnal(range(19, 24), 0.30, 0.65)
    ev_down = np.where(np.isin(hour, range(0, 7)), 0.70, ev_down)
    ev_up = diurnal(range(17, 22), 0.10, 0.30)
    ev_up = np.where(weekday, ev_up, ev_up * 0.7)

    # Heat-pump thermal shift: tracks heating season + early-morning pre-heat.
    hp_season_factor = pd.Series(season).map({"winter": 1.0, "autumn": 0.6, "spring": 0.6, "summer": 0.15}).to_numpy()
    hp_diurnal = diurnal(range(5, 9), 0.25, 0.45)
    hp_up = hp_diurnal * hp_season_factor
    hp_down = (hp_diurnal * 0.8) * hp_season_factor

    # Industrial DR: tied to production hours, reduced on weekends and during
    # the typical Belgian industrial August shutdown.
    august = month == 8
    ind_fast = np.where(weekday & ~august, diurnal(range(6, 22), 0.65, 0.85), 0.10)
    ind_slow = np.where(weekday & ~august, 0.75, 0.15)

    # Commercial HVAC: business hours, weekdays only.
    hvac = np.where(weekday, diurnal(range(8, 19), 0.10, 0.55), 0.05)

    profiles = {
        "PSP_COO_PLATE_TAILLE": (psp_avail.to_numpy(), psp_avail.to_numpy(), 1.0, "S31", "ENTSO-E 2023 UMM outage events for Belgian Hydro Pumped Storage"),
        "BESS_1H": (np.full(N_HOURS, 0.92), np.full(N_HOURS, 0.92), 1.0, "S32", "Technical availability assumption (battery fleet)"),
        "BESS_4H": (np.full(N_HOURS, 0.92), np.full(N_HOURS, 0.92), 1.0, "S32", "Technical availability assumption (battery fleet)"),
        "EV_SMART_CHARGING": (ev_up, ev_down, 1.0, "S33", "Diurnal plug-in pattern (overnight + workplace charging); weekday/weekend split"),
        "HEATPUMP_SHIFT": (hp_up, hp_down, 1.0, "S33", "Heating-season + early-morning pre-heat shift pattern"),
        "INDUSTRIAL_DR_FAST": (ind_fast, ind_fast, 1.0, "S34", "Weekday production-hour availability; reduced to standby during August shutdown/weekends"),
        "INDUSTRIAL_DR_SLOW": (ind_slow, ind_slow, 1.0, "S34", "Weekday production-hour availability; reduced to standby during August shutdown/weekends"),
        "COMMERCIAL_HVAC": (hvac, hvac, 1.0, "S34", "Weekday business-hours availability"),
    }

    rows = []
    for _, fx in flex.iterrows():
        fid = fx["flex_id"]
        if fid not in profiles:
            continue
        up, dn, budget, source_id, note = profiles[fid]
        rows.append(pd.DataFrame({
            "time_id": cal["time_id"],
            "flex_id": fid,
            "availability_pct_up": up,
            "availability_pct_down": dn,
            "energy_budget_pct": budget,
            "source_id": source_id,
            "data_quality": "historical (ENTSO-E 2023 UMM)" if source_id == "S31" else "robust assumption (documented diurnal/seasonal pattern)",
            "notes": note,
        }))
    return pd.concat(rows, ignore_index=True)


# ---------------------------------------------------------------------------
# Static-sheet refreshes
# ---------------------------------------------------------------------------

def refresh_sources(sources: pd.DataFrame) -> pd.DataFrame:
    new_rows = [
        {"source_id": "S25", "source_name": "ENTSO-E Transparency Platform - Total Load 2023",
         "what_it_supports": "02_RepHours full-year hourly demand (D001-D365)",
         "url": "https://transparency.entsoe.eu/load-domain/r2/totalLoadR2/show", "notes": "BE actual load, hourly mean of 15-min values, calendar year 2023"},
        {"source_id": "S26", "source_name": "ENTSO-E Transparency Platform - Generation per type 2023",
         "what_it_supports": "08_RES_CF_Profiles wind/solar capacity factors",
         "url": "https://transparency.entsoe.eu/generation/r2/actualGenerationPerProductionType/show", "notes": "BE actual wind onshore/offshore/solar generation / ENTSO-E installed capacity"},
        {"source_id": "S27", "source_name": "ENTSO-E Transparency Platform - cross-border flows & day-ahead prices 2023",
         "what_it_supports": "05_IntercoProfiles fixed flows and prices",
         "url": "https://transparency.entsoe.eu/transmission-domain/physicalFlow/show", "notes": "Physical flows both directions per border; day-ahead price BE/FR/NL/DE; GB price unavailable via this API, BE price used as proxy"},
        {"source_id": "S28", "source_name": "ENTSO-E day-ahead forecasts vs actuals 2023",
         "what_it_supports": "12_Reserve_ForecastError empirical forecast-error std-dev",
         "url": "https://transparency.entsoe.eu/transmission-domain/r2/dayAheadAggregatedForecastsForGenerationAndLoad/show", "notes": "Per-season std of (actual-forecast)/reference for load, wind, solar; reserve MW volumes remain Elia-informed assumptions (S20/S21)"},
        {"source_id": "S29", "source_name": "ENTSO-E Transparency Platform - Unavailability of Generation Units (UMM/REMIT) 2023",
         "what_it_supports": "11_Availability_Outages nuclear/gas/biomass availability profiles",
         "url": "https://transparency.entsoe.eu/outage-domain/r2/unavailabilityOfGenerationUnits/show", "notes": "Only REMIT-reporting units (>=100MW) are covered; OCGT/CHP/hydro-run units below threshold use the same fleet-level Fossil Gas profile or a seasonal assumption"},
        {"source_id": "S30", "source_name": "Seasonal maintenance assumption",
         "what_it_supports": "11_Availability_Outages for CHP and run-of-river hydro (no UMM coverage)",
         "url": "", "notes": "documented in docs/METHODOLOGY.md (§6) item 7"},
        {"source_id": "S31", "source_name": "ENTSO-E Transparency Platform - UMM Hydro Pumped Storage 2023",
         "what_it_supports": "10_FlexAvailability PSP_COO_PLATE_TAILLE availability",
         "url": "https://transparency.entsoe.eu/outage-domain/r2/unavailabilityOfGenerationUnits/show", "notes": "Coo/Plate-Taille pumped-storage outage events"},
        {"source_id": "S32", "source_name": "Battery storage technical availability assumption",
         "what_it_supports": "10_FlexAvailability BESS_1H/BESS_4H",
         "url": "", "notes": "Asset-level BESS availability is not published (restricted, item 5 in data-source map); flat 92% technical-availability assumption"},
        {"source_id": "S33", "source_name": "EV/heat-pump diurnal+seasonal flexibility assumption",
         "what_it_supports": "10_FlexAvailability EV_SMART_CHARGING / HEATPUMP_SHIFT",
         "url": "https://arxiv.org/abs/2409.18105", "notes": "Documented diurnal plug-in / heating-season pattern, no asset-level metering available (item 14 in data-source map)"},
        {"source_id": "S34", "source_name": "Industrial/commercial DR availability assumption",
         "what_it_supports": "10_FlexAvailability INDUSTRIAL_DR_FAST/SLOW, COMMERCIAL_HVAC",
         "url": "", "notes": "Weekday production/business-hours pattern with August industrial shutdown; aggregator-level data is restricted (item 14 in data-source map)"},
        {"source_id": "S35", "source_name": "Belgium FNA data-source map",
         "what_it_supports": "All sheets - canonical mapping of input categories to Belgium/EU sources",
         "url": "", "notes": "See docs/METHODOLOGY.md (§6)"},
    ]
    return pd.concat([sources, pd.DataFrame(new_rows)], ignore_index=True)


def refresh_res_portfolios(res: pd.DataFrame) -> pd.DataFrame:
    res = res.copy()
    # PV grew from ~9.1 GW (2023/2024) toward ~11.6 GW by 2025 (S12); scale the
    # 2025 PV rows proportionally and refresh provenance.
    pv_scale = 11600.0 / 9130.0
    pv_mask = res["technology"] == "solar_pv"
    res.loc[pv_mask, "capacity_MW_2025"] = (res.loc[pv_mask, "capacity_MW_2025"] * pv_scale).round(0)
    res.loc[pv_mask, "source_id"] = "S12/S26"
    res.loc[pv_mask, "data_quality"] = "scaled from ENTSO-E 2023 actual (6475 MW) toward S12 2025 PV estimate (11.6 GW)"

    wind_mask = res["technology"].isin(["offshore_wind", "onshore_wind"])
    res.loc[wind_mask, "source_id"] = res.loc[wind_mask, "source_id"] + "/S26"
    res.loc[wind_mask, "notes"] = res.loc[wind_mask, "notes"].astype(str) + " Cross-checked against ENTSO-E 2023 installed capacity (offshore 2262 MW, onshore 3053 MW)."
    return res


def refresh_flex_storage(flex: pd.DataFrame) -> pd.DataFrame:
    flex = flex.copy()
    mask = flex["flex_id"] == "PSP_COO_PLATE_TAILLE"
    flex.loc[mask, ["up_capacity_MW_2025", "down_capacity_MW_2025", "up_capacity_MW_2030", "down_capacity_MW_2030"]] = INSTALLED["hydro_psp"]
    flex.loc[mask, "source_id"] = "S31"
    flex.loc[mask, "data_quality"] = "ENTSO-E 2023 installed Hydro Pumped Storage capacity (1308 MW)"
    return flex


def refresh_control(control: pd.DataFrame) -> pd.DataFrame:
    control = control.copy()
    updates = {
        "time_resolution": ("full_year_hourly_8760_2023base", "category", "365 calendar days x 24h, no clustering (D001..D365)", "model design", "S25"),
        "entsoe_data_year": (2023, "year", "ENTSO-E data year used to build the full-year hourly profiles", "user input", "S25"),
        "scenario_id": ("BE_FNA_v3.1_FULLYEAR2023", "Text", "Full-year hourly Belgium FNA input set (8760 h, 2023 base year)", "assumption", "S0"),
    }
    for param, (value, unit, desc, dq, src) in updates.items():
        mask = control["parameter"] == param
        if mask.any():
            control.loc[mask, "value"] = value
            control.loc[mask, "unit"] = unit
            control.loc[mask, "description"] = desc
            control.loc[mask, "data_quality"] = dq
            control.loc[mask, "source_id"] = src
        else:
            new_row = {c: np.nan for c in control.columns}
            new_row.update({"parameter": param, "value": value, "unit": unit, "description": desc, "data_quality": dq, "source_id": src})
            control = pd.concat([control, pd.DataFrame([new_row])], ignore_index=True)
    return control


def refresh_readme(readme: pd.DataFrame) -> pd.DataFrame:
    extra = pd.DataFrame([
        {readme.columns[0]: "Full-year variant", readme.columns[1]:
         "This workbook (v3.1 FullYear2023) replaces the 10 clustered representative days with all 365 days of "
         "2023 (8760 hourly rows, D001..D365, weight_days=1 each). Demand, wind/solar capacity factors, "
         "cross-border flows, day-ahead prices, forecast-error stats and generation outages are ENTSO-E 2023 "
         "actuals; flexible-resource availability and reserve dimensioning remain documented assumptions per "
         "docs/METHODOLOGY.md (§6). Known limitation: storage SOC cycling is still enforced per "
         "calendar day (same_day_pairs in io_excel.write_inc_files), not per year - multi-day storage dynamics "
         "are therefore not captured without a GAMS-side change."},
    ])
    return pd.concat([readme, extra], ignore_index=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Copying {SRC_WB} -> {OUT_WB}")
    shutil.copy(SRC_WB, OUT_WB)

    print("Loading cached ENTSO-E series...")
    load = load_aligned("load")
    res_actual = load_aligned("res_generation_actual")
    res_fc = load_aligned("res_generation_forecast")
    prices = load_aligned("day_ahead_prices")
    flows = load_aligned("cross_border_flows")
    nb_prices = load_aligned("neighbour_day_ahead_prices")

    cal = build_calendar()

    print("Reading static input sheets...")
    res_portfolios = pd.read_excel(SRC_WB, sheet_name="07_RES_Portfolios")
    flex = pd.read_excel(SRC_WB, sheet_name="09_FlexStorage")
    dispatchable = pd.read_excel(SRC_WB, sheet_name="06_DispatchableBlocks")
    borders = pd.read_excel(SRC_WB, sheet_name="04_Interconnectors")
    sources = pd.read_excel(SRC_WB, sheet_name="20_Sources")
    control = pd.read_excel(SRC_WB, sheet_name="01_Control")
    readme = pd.read_excel(SRC_WB, sheet_name="00_ReadMe")

    print("Building 02_RepHours / 03_RepDays (8760 h, 365 days)...")
    hours, days = build_hours_and_days(cal, load)

    print("Building 08_RES_CF_Profiles...")
    res_cf = build_res_cf(cal, res_actual, res_portfolios)

    print("Building 05_IntercoProfiles...")
    interco = build_interco_profiles(cal, flows, prices, nb_prices, borders)

    print("Building 12_Reserve_ForecastError...")
    reserve = build_reserve_forecast_error(cal, load, res_actual, res_fc)

    print("Building 11_Availability_Outages...")
    outages = build_availability_outages(cal, dispatchable)

    print("Building 10_FlexAvailability...")
    flex_avail = build_flex_availability(cal, flex)

    print("Refreshing static sheets...")
    res_portfolios_new = refresh_res_portfolios(res_portfolios)
    flex_new = refresh_flex_storage(flex)
    sources_new = refresh_sources(sources)
    control_new = refresh_control(control)
    readme_new = refresh_readme(readme)

    sheets = {
        "02_RepHours": hours,
        "03_RepDays": days,
        "08_RES_CF_Profiles": res_cf,
        "05_IntercoProfiles": interco,
        "12_Reserve_ForecastError": reserve,
        "11_Availability_Outages": outages,
        "10_FlexAvailability": flex_avail,
        "07_RES_Portfolios": res_portfolios_new,
        "09_FlexStorage": flex_new,
        "20_Sources": sources_new,
        "01_Control": control_new,
        "00_ReadMe": readme_new,
    }

    print("Writing sheets...")
    with pd.ExcelWriter(OUT_WB, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)
            print(f"  {name}: {df.shape}")

    print("Applying header styling and sheet order...")
    wb = openpyxl.load_workbook(OUT_WB)
    for name in sheets:
        header_style(wb[name])
    wb._sheets.sort(key=lambda sheet: sheet_sort_key(sheet.title))
    wb.save(OUT_WB)

    print(f"Done. Wrote {OUT_WB}")


if __name__ == "__main__":
    main()
