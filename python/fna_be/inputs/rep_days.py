"""
Create representative days from ENTSO-E data and write them to the input workbook.

The script updates the same input sheets used by the GAMS workflow:
    - 02_RepHours
    - 03_RepDays
    - 08_RES_CF_Profiles

Configuration is read from the existing project config/workbook first, with
environment variables available for values that should not be stored in code.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

from fna_be.config import (
    ENTSOE_API_KEY,
    ENTSOE_BORDER_TO_COUNTRY,
    ENTSOE_COUNTRY_CODE,
    ENTSOE_DATA_YEAR,
    ENTSOE_NEIGHBOURS,
    ENTSOE_TIMEZONE,
    ENTSOE_WIND_FALLBACK_SHARE,
    EXCEL_FILENAME,
    PROJECT_ROOT,
    REPRESENTATIVE_DAY_AVAILABILITY_PCT,
    REPRESENTATIVE_DAY_CLUSTERS,
    REPRESENTATIVE_DAY_RANDOM_SEED,
    REPRESENTATIVE_DAY_SHEETS,
    REPRESENTATIVE_DAY_SOURCE_ID,
    REPRESENTATIVE_DAY_TARGET_DAYS,
)
from fna_be.io.excel import ensure_workbook_is_not_open, sort_openpyxl_sheets


WORKBOOK_SHEETS = REPRESENTATIVE_DAY_SHEETS
SHEETS_TO_UPDATE = {
    "hours": WORKBOOK_SHEETS["hours"],
    "days": WORKBOOK_SHEETS["days"],
    "res_profiles": WORKBOOK_SHEETS["res_profiles"],
}


@dataclass(frozen=True)
class RepresentativeDayConfig:
    """Runtime settings resolved from config.py, Excel, and environment."""

    api_key: str
    excel_path: Path
    data_year: int
    target_year: int
    country_code: str
    neighbours: list[str]
    n_clusters: int
    random_seed: int
    availability_pct: float
    wind_fallback_share: float
    source_id: str
    timezone: str
    target_days: float


def main() -> None:
    cfg = load_config()
    ensure_workbook_is_not_open(cfg.excel_path)
    print(f"Fetching ENTSO-E data for {cfg.country_code}, {cfg.data_year}...")

    try:
        from entsoe import EntsoePandasClient
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency 'entsoe-py'. Install it before fetching ENTSO-E data.") from exc

    client = EntsoePandasClient(api_key=cfg.api_key)
    profiles = fetch_hourly_profiles(client, cfg)
    representative = cluster_representative_days(
        profiles, cfg.n_clusters, cfg.random_seed, cfg.target_days
    )
    workbook_tables = read_workbook_tables(cfg.excel_path)

    rep_hours_df = build_rep_hours(profiles, representative, cfg)
    rep_days_df = build_rep_days(representative, cfg.source_id)
    res_cf_df = build_res_cf_profiles(profiles, representative, workbook_tables["res"], cfg)
    validate_output_tables(rep_hours_df, rep_days_df, res_cf_df, cfg)

    write_representative_tables(
        cfg.excel_path,
        {
            SHEETS_TO_UPDATE["hours"]: rep_hours_df,
            SHEETS_TO_UPDATE["days"]: rep_days_df,
            SHEETS_TO_UPDATE["res_profiles"]: res_cf_df,
        },
    )

    print(f"Done. Updated {cfg.excel_path} with {len(representative.days)} representative days.")
    for i, weight in enumerate(representative.weights, start=1):
        print(f"  RD{i:02d}: {weight:.1f} days")


def load_config() -> RepresentativeDayConfig:
    """Resolve settings without hard-coding project-specific run choices."""

    excel_path = PROJECT_ROOT / "excel" / EXCEL_FILENAME
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")

    tables = read_workbook_tables(excel_path)
    control = control_dict(tables["control"])
    target_year = int(float(control.get("target_year", 2025)))
    data_year = ENTSOE_DATA_YEAR or int(float(control.get("entsoe_data_year", target_year)))
    n_clusters = REPRESENTATIVE_DAY_CLUSTERS or len(tables["days"])
    country_code = str(control.get("entsoe_country_code", ENTSOE_COUNTRY_CODE)).strip().upper()
    neighbours = ENTSOE_NEIGHBOURS or infer_neighbours(tables["borders"], country_code)

    api_key = ENTSOE_API_KEY.strip()
    if not api_key:
        raise RuntimeError("ENTSOE_API_KEY is empty in config.py.")

    return RepresentativeDayConfig(
        api_key=api_key,
        excel_path=excel_path,
        data_year=data_year,
        target_year=target_year,
        country_code=country_code,
        neighbours=neighbours,
        n_clusters=n_clusters,
        random_seed=REPRESENTATIVE_DAY_RANDOM_SEED,
        availability_pct=REPRESENTATIVE_DAY_AVAILABILITY_PCT,
        wind_fallback_share=ENTSOE_WIND_FALLBACK_SHARE,
        source_id=REPRESENTATIVE_DAY_SOURCE_ID,
        timezone=ENTSOE_TIMEZONE,
        target_days=REPRESENTATIVE_DAY_TARGET_DAYS,
    )


def read_workbook_tables(excel_path: Path) -> dict[str, pd.DataFrame]:
    """Read the workbook sheets needed for configuration and profile output."""

    return {
        "control": read_excel_sheet(excel_path, WORKBOOK_SHEETS["control"]),
        "days": read_excel_sheet(excel_path, WORKBOOK_SHEETS["days"]),
        "borders": read_excel_sheet(excel_path, WORKBOOK_SHEETS["borders"]),
        "res": read_excel_sheet(excel_path, WORKBOOK_SHEETS["res"]),
    }


def read_excel_sheet(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df.columns = [clean_col(c) for c in df.columns]
    return df.dropna(how="all")


def clean_col(value: Any) -> str:
    """Match the column normalization used by the main Excel input reader."""

    return str(value).strip().replace(" ", "_").replace("-", "_").lower()


def control_dict(control_df: pd.DataFrame) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for _, row in control_df.iterrows():
        key = str(row.get("parameter", "")).strip().lower()
        if key:
            out[key] = row.get("value")
    return out


def infer_neighbours(borders: pd.DataFrame, country_code: str) -> list[str]:
    """Map workbook interconnector rows such as BE_FR to ENTSO-E country codes."""

    neighbours: list[str] = []
    for _, row in borders.iterrows():
        candidates = [row.get("border_id"), row.get("border")]
        for value in candidates:
            code = parse_neighbour_code(value, country_code)
            if code and code not in neighbours:
                neighbours.append(code)
                break
    return neighbours


def parse_neighbour_code(value: Any, country_code: str) -> str | None:
    if value is None or pd.isna(value):
        return None
    raw = str(value).strip().upper().replace(" ", "_").replace("-", "_")
    if "_" in raw:
        parts = [part for part in raw.split("_") if part and part != country_code]
        raw = parts[-1] if parts else raw
    return ENTSOE_BORDER_TO_COUNTRY.get(raw)


def fetch_hourly_profiles(client: Any, cfg: RepresentativeDayConfig) -> pd.DataFrame:
    """Fetch load, RES, and cross-border data on a common hourly index."""

    start = pd.Timestamp(f"{cfg.data_year}-01-01", tz=cfg.timezone)
    end = pd.Timestamp(f"{cfg.data_year}-12-31 23:00", tz=cfg.timezone)

    load = hourly_series(client.query_load(cfg.country_code, start=start, end=end), "load")
    wind, solar, source_note = fetch_res_generation(client, cfg, start, end)
    net_import = fetch_net_import(client, cfg, load.index, start, end)

    df = pd.DataFrame(
        {
            "load": load,
            "wind": wind.reindex(load.index),
            "solar": solar.reindex(load.index),
            "net_import": net_import.reindex(load.index),
        }
    )
    df["res_source_note"] = source_note
    add_optional_forecast_errors(client, cfg, df, start, end)

    # Daylight-saving transitions create 23/25-hour local days. The clustering
    # step below keeps only complete 24-hour days so the GAMS time sets stay tidy.
    df.index = df.index.tz_localize(None)
    numeric_cols = [c for c in df.columns if c != "res_source_note"]
    df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    print(f"Built {len(df)} hourly profile rows from ENTSO-E.")
    return df


def fetch_res_generation(
    client: Any,
    cfg: RepresentativeDayConfig,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> tuple[pd.Series, pd.Series, str]:
    """Fetch separate wind/solar generation, with a documented forecast fallback."""

    try:
        wind_parts = [
            series
            for series in (
                fetch_generation_by_psr(client, cfg.country_code, start, end, "B18", "wind_offshore"),
                fetch_generation_by_psr(client, cfg.country_code, start, end, "B19", "wind_onshore"),
            )
            if series is not None
        ]
        if not wind_parts:
            raise ValueError("No offshore or onshore wind generation returned")
        wind = sum_aligned_series(wind_parts, "wind")
        wind.name = "wind"
        solar = fetch_generation_by_psr(client, cfg.country_code, start, end, "B16", "solar")
        if solar is None:
            raise ValueError("No solar generation returned")
        print("Fetched separate wind and solar generation.")
        return wind, solar, "ENTSO-E generation by technology"
    except Exception as exc:
        print(f"Separate wind/solar generation unavailable: {exc}")
        print("Using ENTSO-E wind and solar forecast with configurable split.")
        combined = hourly_series(
            client.query_wind_and_solar_forecast(cfg.country_code, start=start, end=end),
            "wind_solar",
        )
        wind_share = min(max(cfg.wind_fallback_share, 0.0), 1.0)
        return (
            combined * wind_share,
            combined * (1.0 - wind_share),
            "ENTSO-E wind+solar forecast split into wind/solar",
        )


def fetch_generation_by_psr(
    client: Any,
    country_code: str,
    start: pd.Timestamp,
    end: pd.Timestamp,
    psr_type: str,
    name: str,
) -> pd.Series | None:
    """Fetch one ENTSO-E generation PSR type and return None if unavailable."""

    try:
        return hourly_series(
            client.query_generation(country_code, start=start, end=end, psr_type=psr_type),
            name,
        )
    except Exception as exc:
        print(f"Skipped generation {name} ({psr_type}): {exc}")
        return None


def sum_aligned_series(series_list: list[pd.Series], name: str) -> pd.Series:
    """Sum time series with possibly different indexes."""

    total = series_list[0].copy()
    for series in series_list[1:]:
        total = total.add(series, fill_value=0.0)
    total.name = name
    return total


def fetch_net_import(
    client: Any,
    cfg: RepresentativeDayConfig,
    index: pd.DatetimeIndex,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> pd.Series:
    """Sum available border flows into a single net-import clustering feature."""

    net_import = pd.Series(0.0, index=index)
    for neighbour in cfg.neighbours:
        try:
            flows = hourly_series(
                client.query_crossborder_flows(cfg.country_code, neighbour, start=start, end=end),
                f"flow_{neighbour}",
            )
            net_import = net_import.add(flows.reindex(index), fill_value=0.0)
        except Exception as exc:
            print(f"Skipped {cfg.country_code}-{neighbour} flow: {exc}")
    return net_import


def add_optional_forecast_errors(
    client: Any,
    cfg: RepresentativeDayConfig,
    df: pd.DataFrame,
    start: pd.Timestamp,
    end: pd.Timestamp,
) -> None:
    """Add forecast-error features when ENTSO-E exposes the matching series."""

    optional_queries = [
        ("load_forecast_error", "query_load_forecast", "load"),
        ("wind_forecast_error", "query_wind_forecast", "wind"),
        ("solar_forecast_error", "query_solar_forecast", "solar"),
    ]
    for column, query_name, actual_column in optional_queries:
        query = getattr(client, query_name, None)
        if query is None:
            continue
        try:
            forecast = hourly_series(query(cfg.country_code, start=start, end=end), column)
            df[column] = df[actual_column] - forecast.reindex(df.index)
        except Exception:
            continue


def hourly_series(data: Any, name: str) -> pd.Series:
    """Normalize ENTSO-E Series/DataFrame responses to one hourly numeric Series."""

    if isinstance(data, pd.DataFrame):
        series = data.select_dtypes(include="number")
        series = series.iloc[:, 0] if not series.empty else data.iloc[:, 0]
    else:
        series = data
    series = pd.Series(series).resample("1h").mean().squeeze()
    series.name = name
    return pd.to_numeric(series, errors="coerce")


@dataclass(frozen=True)
class RepresentativeDays:
    days: list[Any]
    weights: list[float]
    labels: np.ndarray
    feature_dates: list[Any]


def cluster_representative_days(
    df: pd.DataFrame, n_clusters: int, random_seed: int, target_days: float = 365.0
) -> RepresentativeDays:
    """Cluster complete 24-hour days and pick the real day closest to each centroid.

    Weights are the cluster sizes rescaled so they sum to `target_days`. This
    corrects the bias from dropping daylight-saving-irregular days: the ~362
    complete days are stretched back to a full 365-day year so weighted annual
    demand and energy are not understated.
    """

    try:
        from sklearn.cluster import KMeans
    except ModuleNotFoundError as exc:
        raise RuntimeError("Missing dependency 'scikit-learn'. Install it before clustering representative days.") from exc

    feature_columns = [c for c in df.columns if c != "res_source_note"]
    dates: list[Any] = []
    features: list[np.ndarray] = []

    for date, day_df in df.groupby(df.index.date):
        if len(day_df) != 24:
            continue
        features.append(day_df[feature_columns].to_numpy().reshape(-1))
        dates.append(date)

    if not features:
        raise ValueError("No complete 24-hour days found for clustering.")
    if n_clusters > len(features):
        raise ValueError(f"Requested {n_clusters} clusters but only found {len(features)} complete days.")

    x = np.array(features)
    print(f"Clustering {len(x)} complete days into {n_clusters} representative days...")
    kmeans = KMeans(n_clusters=n_clusters, random_state=random_seed, n_init=10)
    labels = kmeans.fit_predict(x)

    representative_days: list[Any] = []
    raw_counts: list[float] = []
    for cluster_id in range(n_clusters):
        cluster_indices = np.where(labels == cluster_id)[0]
        centroid = kmeans.cluster_centers_[cluster_id]
        distances = np.linalg.norm(x[cluster_indices] - centroid, axis=1)
        best_idx = cluster_indices[int(np.argmin(distances))]
        representative_days.append(dates[best_idx])
        raw_counts.append(float(len(cluster_indices)))

    complete_days = sum(raw_counts)
    scale = float(target_days) / complete_days if complete_days > 0 else 1.0
    weights = [count * scale for count in raw_counts]
    print(
        f"Clustered {int(complete_days)} complete days; weights rescaled to "
        f"{sum(weights):.1f} day-equivalents (target {target_days:.0f})."
    )
    return RepresentativeDays(representative_days, weights, labels, dates)


def build_rep_hours(
    profiles: pd.DataFrame,
    representative: RepresentativeDays,
    cfg: RepresentativeDayConfig,
) -> pd.DataFrame:
    """Build sheet 02_RepHours using the existing workbook column contract."""

    rows: list[dict[str, Any]] = []
    for day_number, rep_date in enumerate(representative.days, start=1):
        rep_day_id = f"RD{day_number:02d}"
        day_df = profiles[profiles.index.date == rep_date].sort_index()
        for hour in range(24):
            row = day_df.iloc[hour]
            time_id = f"{rep_day_id}_H{hour:02d}"
            rows.append(
                {
                    "time_id": time_id,
                    "rep_day_id": rep_day_id,
                    "hour": hour,
                    "next_time_id": f"{rep_day_id}_H{(hour + 1) % 24:02d}",
                    "season": season_from_month(rep_date.month),
                    "day_type": day_type_from_date(rep_date),
                    "weight_days": representative.weights[day_number - 1],
                    "weight_hours": representative.weights[day_number - 1],
                    "chronology_group": rep_day_id,
                    f"gross_demand_MW_{cfg.target_year}": row["load"],
                    "source_id": cfg.source_id,
                    "data_quality": f"ENTSO-E {cfg.data_year}",
                    "notes": f"Representative date {rep_date}; cluster size {representative.weights[day_number - 1]:.0f} days",
                }
            )
    return pd.DataFrame(rows)


def build_rep_days(representative: RepresentativeDays, source_id: str | None = None) -> pd.DataFrame:
    """Build sheet 03_RepDays from clustering weights."""

    total_days = float(sum(representative.weights))
    return pd.DataFrame(
        {
            "rep_day_id": [f"RD{i:02d}" for i in range(1, len(representative.days) + 1)],
            "description": [f"Cluster {i:02d} representative date {d}" for i, d in enumerate(representative.days, start=1)],
            "season": [season_from_month(d.month) for d in representative.days],
            "day_type": [day_type_from_date(d) for d in representative.days],
            "weight_days": representative.weights,
            "selection_reason": [
                f"Closest complete day to centroid of {weight:.0f} clustered days"
                for weight in representative.weights
            ],
            "probability_pct": [weight / total_days * 100.0 for weight in representative.weights],
            "source_id": source_id or REPRESENTATIVE_DAY_SOURCE_ID,
            "data_quality": "ENTSO-E derived",
        }
    )


def build_res_cf_profiles(
    profiles: pd.DataFrame,
    representative: RepresentativeDays,
    res_portfolios: pd.DataFrame,
    cfg: RepresentativeDayConfig,
) -> pd.DataFrame:
    """Build RES capacity factors from generation divided by workbook capacity."""

    capacity_col = first_existing_column(
        res_portfolios,
        [f"capacity_mw_{cfg.target_year}", "capacity_mw"],
    )
    if capacity_col is None:
        raise KeyError(f"No RES capacity column found for target year {cfg.target_year}.")

    rows: list[dict[str, Any]] = []
    active_res = res_portfolios[pd.to_numeric(res_portfolios[capacity_col], errors="coerce").fillna(0.0) > 0.0]

    for day_number, rep_date in enumerate(representative.days, start=1):
        rep_day_id = f"RD{day_number:02d}"
        day_df = profiles[profiles.index.date == rep_date].sort_index()
        for hour in range(24):
            profile_row = day_df.iloc[hour]
            for _, res_row in active_res.iterrows():
                technology = str(res_row.get("technology", "")).lower()
                generation_col = generation_column_for_technology(technology)
                if generation_col is None:
                    continue
                capacity = float(res_row[capacity_col])
                capacity_factor = profile_row[generation_col] / capacity if capacity > 0 else 0.0
                rows.append(
                    {
                        "time_id": f"{rep_day_id}_H{hour:02d}",
                        "res_id": res_row["res_id"],
                        "capacity_factor": float(np.clip(capacity_factor, 0.0, 1.0)),
                        "availability_pct": cfg.availability_pct,
                        "source_id": cfg.source_id,
                        "data_quality": f"ENTSO-E {cfg.data_year}",
                        "notes": f"Representative date {rep_date}; {profiles['res_source_note'].iloc[0]}",
                    }
                )
    return pd.DataFrame(rows)


def validate_output_tables(
    rep_hours_df: pd.DataFrame,
    rep_days_df: pd.DataFrame,
    res_cf_df: pd.DataFrame,
    cfg: RepresentativeDayConfig,
) -> None:
    """Catch empty or malformed outputs before replacing workbook sheets."""

    required = {
        WORKBOOK_SHEETS["hours"]: {
            "time_id",
            "rep_day_id",
            "hour",
            "next_time_id",
            "weight_days",
            f"gross_demand_MW_{cfg.target_year}",
        },
        WORKBOOK_SHEETS["days"]: {"rep_day_id", "weight_days", "probability_pct"},
        WORKBOOK_SHEETS["res_profiles"]: {"time_id", "res_id", "capacity_factor", "availability_pct"},
    }
    tables = {
        WORKBOOK_SHEETS["hours"]: rep_hours_df,
        WORKBOOK_SHEETS["days"]: rep_days_df,
        WORKBOOK_SHEETS["res_profiles"]: res_cf_df,
    }
    for sheet_name, df in tables.items():
        if df.empty:
            raise ValueError(f"Generated table for {sheet_name} is empty; Excel was not updated.")
        missing = required[sheet_name] - set(df.columns)
        if missing:
            raise ValueError(f"Generated table for {sheet_name} is missing columns: {sorted(missing)}")

    if len(rep_days_df) != cfg.n_clusters:
        raise ValueError(f"Expected {cfg.n_clusters} representative days, generated {len(rep_days_df)}.")

    weights = pd.to_numeric(rep_days_df["weight_days"], errors="coerce")
    if weights.isna().any() or (weights <= 0).any():
        raise ValueError("Representative-day weights must be positive numeric values.")

    probability_sum = pd.to_numeric(rep_days_df["probability_pct"], errors="coerce").sum()
    if abs(float(probability_sum) - 100.0) > 1e-6:
        raise ValueError(f"Representative-day probabilities must sum to 100%, got {probability_sum:.8f}%.")

    hour_weights = (
        rep_hours_df[["rep_day_id", "weight_days"]]
        .drop_duplicates()
        .set_index("rep_day_id")["weight_days"]
        .astype(float)
    )
    day_weights = rep_days_df.set_index("rep_day_id")["weight_days"].astype(float)
    if not hour_weights.equals(day_weights):
        raise ValueError("02_RepHours weight_days do not match 03_RepDays weight_days.")


def write_representative_tables(excel_path: Path, tables: dict[str, pd.DataFrame]) -> None:
    """Replace workbook sheets using openpyxl and save the file explicitly."""

    ensure_workbook_is_not_open(excel_path)
    try:
        workbook = load_workbook(excel_path)
        for sheet_name, df in tables.items():
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            worksheet = workbook.create_sheet(sheet_name)
            write_dataframe_to_sheet(worksheet, df)
        sort_openpyxl_sheets(workbook)
        workbook.save(excel_path)
    except PermissionError as exc:
        raise PermissionError(f"Cannot write {excel_path}. Close the workbook in Excel and retry.") from exc


def write_dataframe_to_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Small openpyxl writer mirroring the plain table style used elsewhere."""

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_index, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(df.itertuples(index=False), start=2):
        for col_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=col_index, value=excel_value(value))
    worksheet.freeze_panes = "A2"


def excel_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    return value


def generation_column_for_technology(technology: str) -> str | None:
    if "wind" in technology:
        return "wind"
    if "solar" in technology or "pv" in technology:
        return "solar"
    return None


def first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lower_to_actual = {str(col).lower(): col for col in df.columns}
    for candidate in candidates:
        if candidate.lower() in lower_to_actual:
            return lower_to_actual[candidate.lower()]
    return None


def season_from_month(month: int) -> str:
    if month in (12, 1, 2):
        return "winter"
    if month in (3, 4, 5):
        return "spring"
    if month in (6, 7, 8):
        return "summer"
    return "autumn"


def day_type_from_date(date_value: Any) -> str:
    return "weekday" if pd.Timestamp(date_value).weekday() < 5 else "weekend"


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        raise SystemExit(f"Error: {exc}") from None
