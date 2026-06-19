"""
migrate_v3_1_workbook.py - one-time workbook migration for v3.1 indicators.

Adds the input sheets/columns needed by the v3.1 ACER gap-closing indicators
(DSO/TSO/Article-14 split, unavailability needs, barriers/digitalisation,
ramping-capacity translation):

    - 13_NetworkNeeds: adds `timeframe` (structural/fine_tuning) and `zone_id`
      columns (default: structural / TSO for existing rows).
    - 13b_DSO_Zones: new sheet, local hosting/feeder-capacity proxy inputs.
    - 10b_Prequalification_Log: new sheet, one row per 09_FlexStorage flex_id.
    - 17_Barriers_Digitalisation: new sheet, example barrier register.
    - 01_Control: adds `market_time_unit_minutes` parameter.

Idempotent: re-running on an already-migrated workbook makes no changes.
Close the workbook in Excel before running this script.

Usage:
    python migrate_v3_1_workbook.py [path/to/workbook.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


from fna_be.config import PROJECT_ROOT, EXCEL_FILENAME  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _set_header_style(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _header_row(ws) -> list[str]:
    return [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]


def migrate_network_needs(wb: openpyxl.Workbook) -> None:
    ws = wb["13_NetworkNeeds"]
    header = _header_row(ws)
    n_cols = len(header)

    new_cols = {"timeframe": "structural", "zone_id": "TSO"}
    for col_name, default in new_cols.items():
        if col_name in header:
            continue
        n_cols += 1
        ws.cell(row=1, column=n_cols, value=col_name)
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value in (None, ""):
                continue
            ws.cell(row=row, column=n_cols, value=default)
    _set_header_style(ws, n_cols)
    print(f"13_NetworkNeeds: columns now {_header_row(ws)}")


def add_dso_zones(wb: openpyxl.Workbook) -> None:
    if "13b_DSO_Zones" in wb.sheetnames:
        print("13b_DSO_Zones already exists, skipping")
        return
    ws = wb.create_sheet("13b_DSO_Zones")
    header = [
        "zone_id", "region", "voltage_level", "hosting_capacity_MW",
        "peak_feeder_capacity_MW", "share_of_national_RES_pct",
        "share_of_national_demand_pct", "source_id", "data_quality", "notes",
    ]
    ws.append(header)
    ws.append([
        "DSO_FLANDERS_LV", "Flanders", "LV/MV", 500, 600, 25, 22, "S22",
        "placeholder",
        "Example local hosting/feeder-capacity proxy; replace with DSO DNDP "
        "hosting-capacity studies and feeder peak-load data.",
    ])
    _set_header_style(ws, len(header))
    print("Added 13b_DSO_Zones")


def add_prequalification_log(wb: openpyxl.Workbook) -> None:
    if "10b_Prequalification_Log" in wb.sheetnames:
        print("10b_Prequalification_Log already exists, skipping")
        return

    flex_ws = wb["09_FlexStorage"]
    flex_header = _header_row(flex_ws)
    flex_id_col = flex_header.index("flex_id") + 1
    flex_ids = []
    for row in range(2, flex_ws.max_row + 1):
        val = flex_ws.cell(row=row, column=flex_id_col).value
        if val:
            flex_ids.append(val)

    ws = wb.create_sheet("10b_Prequalification_Log")
    header = ["flex_id", "prequalification_status", "source_id", "data_quality", "notes"]
    ws.append(header)
    for flex_id in flex_ids:
        ws.append([
            flex_id, "qualified", "S0", "assumption",
            "Default: fully prequalified. Set to 'temporary_limit' or "
            "'unavailable' to model prequalification failures or "
            "network-driven derating of this flexibility resource.",
        ])
    _set_header_style(ws, len(header))
    print(f"Added 10b_Prequalification_Log with {len(flex_ids)} rows")


def add_barriers_digitalisation(wb: openpyxl.Workbook) -> None:
    if "17_Barriers_Digitalisation" in wb.sheetnames:
        print("17_Barriers_Digitalisation already exists, skipping")
        return
    ws = wb.create_sheet("17_Barriers_Digitalisation")
    header = [
        "barrier_id", "category", "description", "severity_1to5",
        "digitalisation_dependency", "status", "source_id", "data_quality", "notes",
    ]
    ws.append(header)
    rows = [
        ("BAR_DSR_BASELINE", "demand_response",
         "No standardised consumption-baseline methodology for industrial/commercial DR providers.",
         4, "baseline_methodology", "open", "S23", "placeholder",
         "Limits independent aggregator participation in balancing/redispatch."),
        ("BAR_DYNAMIC_TARIFFS", "retail_tariffs",
         "Dynamic/time-of-use tariffs not yet standard for residential prosumers.",
         3, "smart_meter", "open", "S23", "placeholder",
         "Reduces price signal for EV/heat-pump smart charging."),
        ("BAR_SMART_METER_ROLLOUT", "metering",
         "Smart meter penetration incomplete, limiting near-real-time data for flexibility activation.",
         3, "smart_meter", "open", "S23", "placeholder",
         "Affects data granularity available for short-term need calibration."),
        ("BAR_AGGREGATOR_ACCESS", "market_access",
         "Independent aggregator market access rules for balancing markets still maturing.",
         3, "aggregator_API", "open", "S23", "placeholder",
         "Affects flexibility resource availability assumptions in 10_FlexAvailability."),
        ("BAR_DSO_DATA_SHARING", "data_sharing",
         "DSO hosting-capacity and congestion data not yet published in machine-readable form.",
         4, "open_data_platform", "open", "S23", "placeholder",
         "Blocks calibration of 13b_DSO_Zones from real DNDP data."),
    ]
    for row in rows:
        ws.append(row)
    _set_header_style(ws, len(header))
    print(f"Added 17_Barriers_Digitalisation with {len(rows)} rows")


def add_control_param(wb: openpyxl.Workbook) -> None:
    ws = wb["01_Control"]
    header = _header_row(ws)
    param_col = header.index("parameter") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=param_col).value == "market_time_unit_minutes":
            print("01_Control: market_time_unit_minutes already present, skipping")
            return

    next_row = ws.max_row + 1
    values = ["market_time_unit_minutes", 60, "minutes",
              "ACER market time unit used to translate ramping MW/h stats into "
              "MW-per-MTU (sheet 41b_FNA_Ramping_Capacity)", "model design", "S1"]
    for col, value in enumerate(values, start=1):
        ws.cell(row=next_row, column=col, value=value)
    print("01_Control: added market_time_unit_minutes = 60")


def sheet_sort_key(name: str) -> tuple[int, str]:
    import re
    match = re.match(r"^(\d+)", str(name))
    if match:
        return int(match.group(1)), str(name).lower()
    return 999, str(name).lower()


def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else PROJECT_ROOT / "excel" / EXCEL_FILENAME
    lock_file = path.with_name(f"~${path.name}")
    if lock_file.exists():
        raise SystemExit(f"Workbook appears open in Excel ({lock_file}). Close it and retry.")

    print(f"Migrating {path}")
    wb = openpyxl.load_workbook(path)

    migrate_network_needs(wb)
    add_dso_zones(wb)
    add_prequalification_log(wb)
    add_barriers_digitalisation(wb)
    add_control_param(wb)

    wb._sheets.sort(key=lambda sheet: sheet_sort_key(sheet.title))

    wb.save(path)
    print(f"Saved {path}")


if __name__ == "__main__":
    main()
